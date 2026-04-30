from __future__ import annotations

import argparse
import csv
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parent.parent
MANUAL_DIR = ROOT / "data" / "manual"
DOCS_DIR = ROOT / "docs"

DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COST_TABS = ("Tools", "Parts", "Substances", "Wiring", "Service")
MANAGED_KEEP_TABS = ("Plan_Summary", "Procurement_Pass2", "Master_Costs", "Build_Plan", "Restore_Replace", "Tab_Reconciliation")
MANAGED_DROP_TABS = (
    "Procurement_Baskets",
    "Reassembly_Packages",
    "Reassembly_Dependency",
    "Component_Disposition",
    "Procurement_Base",
    "Cost_Reconciliation",
    "Workbook_Tidy_Backlog",
)
HIDE_REFERENCE_TABS = ("Electrical Templates", "Fabrication")

CSV_PROCUREMENT_PASS2 = MANUAL_DIR / "procurement_decision_matrix_pass2.csv"
CSV_REASSEMBLY_PACKAGES = MANUAL_DIR / "reassembly_work_packages.csv"
CSV_COMPONENT_DISPOSITION = MANUAL_DIR / "component_disposition_plan.csv"
CSV_SHEET_PROFILES = MANUAL_DIR / "j40_workbook_sheet_profiles.csv"
CSV_TIDY_BACKLOG = MANUAL_DIR / "j40_workbook_tidy_backlog.csv"

REPORT_PATH = DOCS_DIR / "workbook-full-reconciliation.md"


@dataclass(frozen=True)
class CostLine:
    source_tab: str
    source_row: int
    item: str
    price_raw: str
    amount: str
    vendor: str
    received_raw: str
    paid_raw: str
    reconciled_status: str
    reconciled_note: str


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return " ".join(text.split())


def parse_amount(text: str) -> str:
    normalized = text.replace(",", "").strip()
    if not normalized:
        return ""
    digits = []
    for ch in normalized:
        if ch.isdigit() or ch in {".", "-"}:
            digits.append(ch)
        elif digits:
            break
    if not digits:
        return ""
    raw = "".join(digits)
    try:
        if "." in raw:
            return str(int(float(raw)))
        return str(int(raw))
    except ValueError:
        return ""


def normalize_flag(text: str) -> str:
    token = clean_text(text).upper()
    if not token:
        return ""
    if token in {"Y", "YES"}:
        return "yes"
    if token in {"N", "NO"}:
        return "no"
    if "COD" in token:
        return "cod"
    if "CANCEL" in token:
        return "cancelled"
    if token == "?":
        return "unknown"
    if "CHECK" in token:
        return "needs_check"
    return "other"


def is_section_header(item: str, price: str, vendor: str, received: str, paid: str) -> bool:
    if not item:
        return False
    return not any([price, vendor, received, paid])


def reconcile_status(item: str, price: str, vendor: str, received: str, paid: str) -> tuple[str, str]:
    if is_section_header(item, price, vendor, received, paid):
        return "section_header", "Label/group row."

    joined = f"{item} {price} {vendor} {received} {paid}".lower()
    received_flag = normalize_flag(received)
    paid_flag = normalize_flag(paid)

    if "re-use" in joined or "reuse" in joined:
        return "reuse_existing", "Marked for re-use, not replacement purchase."
    if received_flag == "cancelled" or paid_flag == "cancelled":
        return "cancelled", "Marked cancelled."
    if received_flag == "yes" or paid_flag in {"yes", "cod"}:
        return "procured_or_on_hand", "Evidence indicates purchased/received."
    if received_flag in {"unknown", "needs_check"} or paid_flag in {"unknown", "needs_check"}:
        return "needs_confirmation", "Status exists but requires confirmation."
    if received_flag == "no" or paid_flag == "no":
        return "not_procured", "Explicitly not purchased/received yet."
    return "planned_or_open", "Open item with no procurement evidence yet."


def autosize_columns(ws: Worksheet, max_width: int = 75) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def style_header(ws: Worksheet, row: int = 1) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def ensure_standard_headers(workbook) -> list[str]:
    notes: list[str] = []

    # Normalize misspelled tab name.
    if "Relay Assigment" in workbook.sheetnames and "Relay Assignment" not in workbook.sheetnames:
        workbook["Relay Assigment"].title = "Relay Assignment"
        notes.append("Renamed tab `Relay Assigment` -> `Relay Assignment`.")

    for tab in ("Tools", "Parts", "Substances", "Wiring"):
        if tab not in workbook.sheetnames:
            continue
        ws = workbook[tab]
        if clean_text(ws["D1"].value).lower() == "receieved":
            ws["D1"] = "Received"
            notes.append(f"Header fixed in `{tab}`: `Receieved` -> `Received`.")
        elif not clean_text(ws["D1"].value):
            ws["D1"] = "Received"
            notes.append(f"Header filled in `{tab}` column D as `Received`.")
        if not clean_text(ws["E1"].value):
            ws["E1"] = "Paid"
            notes.append(f"Header filled in `{tab}` column E as `Paid`.")

    if "Wiring" in workbook.sheetnames:
        ws = workbook["Wiring"]
        if not clean_text(ws["F1"].value):
            ws["F1"] = "Notes"
        if not clean_text(ws["G1"].value):
            ws["G1"] = "Tag"
        notes.append("Standardized `Wiring` headers for columns F/G (`Notes`, `Tag`).")

    if "Bodywork" in workbook.sheetnames:
        ws = workbook["Bodywork"]
        if not clean_text(ws["E1"].value):
            ws["E1"] = "Work Type"
        if not clean_text(ws["F1"].value):
            ws["F1"] = "Notes"
        notes.append("Standardized `Bodywork` headers for columns E/F.")

    if "Service" in workbook.sheetnames:
        ws = workbook["Service"]
        if not clean_text(ws["C1"].value):
            ws["C1"] = "Notes"
        notes.append("Standardized `Service` header for column C (`Notes`).")

    return notes


def split_parts_mixed_tables(workbook) -> list[str]:
    notes: list[str] = []
    if "Parts" not in workbook.sheetnames:
        return notes

    ws = workbook["Parts"]
    planning_header_row = None
    supplier_header_row = None

    for row in range(2, ws.max_row + 1):
        c1 = clean_text(ws.cell(row, 1).value).lower()
        c2 = clean_text(ws.cell(row, 2).value).lower()
        c3 = clean_text(ws.cell(row, 3).value).lower()
        if c1 == "category" and c2 == "subcategory" and c3 == "item":
            planning_header_row = row
        if c1 == "category" and c2 == "best place to start":
            supplier_header_row = row
            break

    if planning_header_row is None:
        return notes

    # Extract replacement estimates block (planning header through row before supplier header, excluding blank-only rows).
    replacement_rows: list[dict[str, str]] = []
    replacement_fields = [
        "category",
        "subcategory",
        "item",
        "qty",
        "source_supplier",
        "estimated_cost_pkr",
        "average_est_cost_pkr",
        "priority",
    ]

    replacement_end = supplier_header_row - 1 if supplier_header_row else ws.max_row
    for row in range(planning_header_row + 1, replacement_end + 1):
        values = [clean_text(ws.cell(row, col).value) for col in range(1, 9)]
        if not any(values):
            continue
        if values[0].lower() == "total":
            continue
        replacement_rows.append(
            {
                "category": values[0],
                "subcategory": values[1],
                "item": values[2],
                "qty": values[3],
                "source_supplier": values[4],
                "estimated_cost_pkr": values[5],
                "average_est_cost_pkr": values[6],
                "priority": values[7],
            }
        )

    replace_table_sheet(workbook, "Parts_Estimates", replacement_rows, replacement_fields)

    # Extract supplier starting points block into a compact sheet.
    supplier_rows: list[dict[str, str]] = []
    supplier_fields = ["category", "best_place_to_start", "why"]
    if supplier_header_row:
        for row in range(supplier_header_row + 1, ws.max_row + 1):
            values = [clean_text(ws.cell(row, col).value) for col in range(1, 4)]
            if not any(values):
                continue
            supplier_rows.append(
                {
                    "category": values[0],
                    "best_place_to_start": values[1],
                    "why": values[2],
                }
            )
    replace_table_sheet(workbook, "Supplier_Starting_Points", supplier_rows, supplier_fields)

    # Keep Parts tab focused on tracked parts records only.
    delete_from = planning_header_row
    ws.delete_rows(delete_from, ws.max_row - delete_from + 1)
    notes.append(
        f"Split mixed planning content out of `Parts`: moved {len(replacement_rows)} estimate rows and {len(supplier_rows)} supplier rows to dedicated tabs."
    )
    return notes


def move_wiring_suspension_block(workbook) -> list[str]:
    notes: list[str] = []
    if "Wiring" not in workbook.sheetnames or "Suspension" not in workbook.sheetnames:
        return notes

    wiring = workbook["Wiring"]
    suspension = workbook["Suspension"]
    suspension_start = None
    component_header = None

    for row in range(2, wiring.max_row + 1):
        first = clean_text(wiring.cell(row, 1).value).lower()
        if first == "suspension":
            suspension_start = row
        if first == "component" and suspension_start is not None:
            component_header = row
            break

    if suspension_start is None or component_header is None:
        return notes

    extracted_rows: list[tuple[str, str, str, str]] = []
    for row in range(component_header + 1, wiring.max_row + 1):
        component = clean_text(wiring.cell(row, 1).value)
        part = clean_text(wiring.cell(row, 2).value)
        quantity = clean_text(wiring.cell(row, 6).value)
        details = clean_text(wiring.cell(row, 7).value)
        if not any([component, part, quantity, details]):
            continue
        extracted_rows.append((component, part, quantity, details))

    # Append unique rows into Suspension tab.
    existing_keys = set()
    for row in range(2, suspension.max_row + 1):
        key = (
            clean_text(suspension.cell(row, 1).value),
            clean_text(suspension.cell(row, 2).value),
            clean_text(suspension.cell(row, 3).value),
            clean_text(suspension.cell(row, 4).value),
        )
        if any(key):
            existing_keys.add(key)

    append_count = 0
    write_row = suspension.max_row + 1
    for entry in extracted_rows:
        if entry in existing_keys:
            continue
        suspension.cell(write_row, 1, entry[0])
        suspension.cell(write_row, 2, entry[1])
        suspension.cell(write_row, 3, entry[2])
        suspension.cell(write_row, 4, entry[3])
        write_row += 1
        append_count += 1

    style_header(suspension)
    autosize_columns(suspension, max_width=65)

    # Remove misplaced suspension block from Wiring.
    wiring.delete_rows(suspension_start, wiring.max_row - suspension_start + 1)
    notes.append(f"Moved suspension block from `Wiring` to `Suspension` (appended {append_count} unique rows).")
    return notes


def dedupe_suspension_tab(workbook) -> list[str]:
    notes: list[str] = []
    if "Suspension" not in workbook.sheetnames:
        return notes

    ws = workbook["Suspension"]
    if ws.max_row < 3:
        return notes

    def score_row(row_idx: int) -> int:
        qty = clean_text(ws.cell(row_idx, 3).value)
        details = clean_text(ws.cell(row_idx, 4).value)
        score = 0
        if qty:
            score += 2
        if details:
            score += 2
        score += 1 if clean_text(ws.cell(row_idx, 2).value) else 0
        score += 1 if clean_text(ws.cell(row_idx, 1).value) else 0
        return score

    groups: defaultdict[tuple[str, str], list[int]] = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        component = clean_text(ws.cell(row, 1).value)
        part = clean_text(ws.cell(row, 2).value)
        qty = clean_text(ws.cell(row, 3).value)
        details = clean_text(ws.cell(row, 4).value)
        if not any([component, part, qty, details]):
            continue
        key = (component.lower(), part.lower())
        groups[key].append(row)

    rows_to_delete: list[int] = []
    for _, row_indexes in groups.items():
        if len(row_indexes) <= 1:
            continue
        ranked = sorted(row_indexes, key=lambda idx: score_row(idx), reverse=True)
        rows_to_delete.extend(ranked[1:])

    if rows_to_delete:
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row, 1)
        notes.append(f"Deduplicated `Suspension` tab: removed {len(rows_to_delete)} lower-detail duplicate rows.")
        style_header(ws)
        autosize_columns(ws, max_width=65)

    return notes


def update_cost_tabs_with_reconciled_columns(workbook) -> tuple[list[CostLine], dict[str, Counter[str]]]:
    lines: list[CostLine] = []
    status_counts: dict[str, Counter[str]] = {}

    for tab in COST_TABS:
        if tab not in workbook.sheetnames:
            continue
        ws = workbook[tab]

        header_map = {clean_text(ws.cell(1, c).value).lower(): c for c in range(1, ws.max_column + 1)}
        item_col = header_map.get("item") or header_map.get("service") or 1
        price_col = header_map.get("price") or header_map.get("cost") or 2
        vendor_col = header_map.get("vendor") or 3
        rec_col = header_map.get("received") or 4
        paid_col = header_map.get("paid") or 5

        reconciled_status_col = None
        reconciled_note_col = None
        for col in range(1, ws.max_column + 1):
            header = clean_text(ws.cell(1, col).value).lower()
            if header == "reconciled_status":
                reconciled_status_col = col
            if header == "reconciled_note":
                reconciled_note_col = col
        if reconciled_status_col is None:
            reconciled_status_col = ws.max_column + 1
            ws.cell(1, reconciled_status_col, "Reconciled_Status")
        if reconciled_note_col is None:
            reconciled_note_col = ws.max_column + 1
            ws.cell(1, reconciled_note_col, "Reconciled_Note")

        if tab == "Parts":
            procured_mark_col = None
            for col in range(1, ws.max_column + 1):
                if clean_text(ws.cell(1, col).value).lower() == "procured_mark":
                    procured_mark_col = col
                    break
            if procured_mark_col is None:
                procured_mark_col = ws.max_column + 1
                ws.cell(1, procured_mark_col, "Procured_Mark")
        else:
            procured_mark_col = None

        counter: Counter[str] = Counter()
        for row in range(2, ws.max_row + 1):
            item = clean_text(ws.cell(row, item_col).value)
            price = clean_text(ws.cell(row, price_col).value)
            vendor = clean_text(ws.cell(row, vendor_col).value)
            received = clean_text(ws.cell(row, rec_col).value)
            paid = clean_text(ws.cell(row, paid_col).value)

            if not any([item, price, vendor, received, paid]):
                continue

            status, note = reconcile_status(item, price, vendor, received, paid)
            ws.cell(row, reconciled_status_col, status)
            ws.cell(row, reconciled_note_col, note)

            if procured_mark_col is not None:
                if status in {"procured_or_on_hand", "reuse_existing"}:
                    ws.cell(row, procured_mark_col, "✓")
                elif status == "cancelled":
                    ws.cell(row, procured_mark_col, "cancelled")
                else:
                    ws.cell(row, procured_mark_col, "")

            counter[status] += 1
            lines.append(
                CostLine(
                    source_tab=tab,
                    source_row=row,
                    item=item,
                    price_raw=price,
                    amount=parse_amount(price),
                    vendor=vendor,
                    received_raw=received,
                    paid_raw=paid,
                    reconciled_status=status,
                    reconciled_note=note,
                )
            )

        status_counts[tab] = counter
        style_header(ws)
        autosize_columns(ws, max_width=70)

    return lines, status_counts


def replace_table_sheet(workbook, sheet_name: str, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(title=sheet_name)
    for idx, header in enumerate(fieldnames, start=1):
        ws.cell(1, idx, header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(fieldnames, start=1):
            ws.cell(row_idx, col_idx, row.get(header, ""))
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{max(2, len(rows) + 1)}"
    autosize_columns(ws, max_width=75)


def build_plan_summary_sheet(workbook, pass2_rows: list[dict[str, str]], status_counts: dict[str, Counter[str]]) -> None:
    if "Plan_Summary" in workbook.sheetnames:
        del workbook["Plan_Summary"]
    ws = workbook.create_sheet(title="Plan_Summary")

    decision_counts = Counter(row.get("pass2_decision", "") for row in pass2_rows)
    timing_counts = Counter(row.get("timing_window", "") for row in pass2_rows)
    immediate = [row for row in pass2_rows if row.get("timing_window") in {"tub_off_immediate", "in_flight_now"}]

    ws["A1"] = "J40 Workbook Reconciled Control Summary"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Workbook"
    ws["B4"] = str(DEFAULT_WORKBOOK_PATH)
    ws["A5"] = "Pass2 rows"
    ws["B5"] = len(pass2_rows)

    ws["A7"] = "Pass2 Decision Counts"
    ws["A7"].font = Font(bold=True)
    r = 8
    for key, value in sorted(decision_counts.items()):
        ws.cell(r, 1, key)
        ws.cell(r, 2, value)
        r += 1

    ws["D7"] = "Timing Windows"
    ws["D7"].font = Font(bold=True)
    r = 8
    for key, value in sorted(timing_counts.items()):
        ws.cell(r, 4, key)
        ws.cell(r, 5, value)
        r += 1

    ws["A18"] = "Immediate Tub-Off Actions"
    ws["A18"].font = Font(bold=True)
    headers = ["entry_id", "item", "pass2_decision", "timing_window"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(19, idx, header)
    style_header(ws, row=19)

    r = 20
    for row in immediate:
        ws.cell(r, 1, row.get("entry_id", ""))
        ws.cell(r, 2, row.get("item", ""))
        ws.cell(r, 3, row.get("pass2_decision", ""))
        ws.cell(r, 4, row.get("timing_window", ""))
        r += 1

    ws["A30"] = "Cost Tab Reconciliation Counts"
    ws["A30"].font = Font(bold=True)
    ws["A31"] = "tab"
    ws["B31"] = "status"
    ws["C31"] = "count"
    style_header(ws, row=31)
    r = 32
    for tab in COST_TABS:
        for status, count in sorted(status_counts.get(tab, Counter()).items()):
            ws.cell(r, 1, tab)
            ws.cell(r, 2, status)
            ws.cell(r, 3, count)
            r += 1

    ws.freeze_panes = "A20"
    autosize_columns(ws)


def build_master_costs_sheet(workbook, lines: list[CostLine]) -> None:
    rows = [
        {
            "source_tab": line.source_tab,
            "source_row": str(line.source_row),
            "item": line.item,
            "price_raw": line.price_raw,
            "amount": line.amount,
            "vendor": line.vendor,
            "received_raw": line.received_raw,
            "paid_raw": line.paid_raw,
            "reconciled_status": line.reconciled_status,
            "reconciled_note": line.reconciled_note,
        }
        for line in lines
    ]
    replace_table_sheet(
        workbook,
        "Master_Costs",
        rows,
        [
            "source_tab",
            "source_row",
            "item",
            "price_raw",
            "amount",
            "vendor",
            "received_raw",
            "paid_raw",
            "reconciled_status",
            "reconciled_note",
        ],
    )


def build_build_plan_sheet(workbook, reassembly_rows: list[dict[str, str]], pass2_rows: list[dict[str, str]]) -> None:
    rows: list[dict[str, str]] = []
    for row in reassembly_rows:
        rows.append(
            {
                "record_type": "work_package",
                "id": row.get("work_package_id", ""),
                "title_or_item": row.get("title", ""),
                "lane_or_workstream": row.get("lane", ""),
                "state_or_decision": row.get("current_state", ""),
                "timing_or_depends_on": row.get("depends_on", ""),
                "action_or_gate": row.get("gate_to_close", ""),
                "notes": row.get("key_procurement_actions", ""),
            }
        )

    for row in pass2_rows:
        if row.get("timing_window") in {"tub_off_immediate", "in_flight_now", "post_tub_off_inspection", "pre_order_audit"}:
            rows.append(
                {
                    "record_type": "procurement",
                    "id": row.get("entry_id", ""),
                    "title_or_item": row.get("item", ""),
                    "lane_or_workstream": row.get("workstream", ""),
                    "state_or_decision": row.get("pass2_decision", ""),
                    "timing_or_depends_on": row.get("timing_window", ""),
                    "action_or_gate": row.get("budget_mode", ""),
                    "notes": row.get("rationale", ""),
                }
            )

    replace_table_sheet(
        workbook,
        "Build_Plan",
        rows,
        [
            "record_type",
            "id",
            "title_or_item",
            "lane_or_workstream",
            "state_or_decision",
            "timing_or_depends_on",
            "action_or_gate",
            "notes",
        ],
    )


def build_restore_replace_sheet(workbook, component_rows: list[dict[str, str]], pass2_rows: list[dict[str, str]]) -> None:
    rows: list[dict[str, str]] = []

    for row in component_rows:
        rows.append(
            {
                "record_type": "existing_component",
                "id": row.get("component_job_id", ""),
                "subject": row.get("component_group", ""),
                "strategy": row.get("recommended_disposition", ""),
                "decision": row.get("reuse_decision", ""),
                "timing_or_gate": row.get("dependency_lane", ""),
                "action": row.get("pre_reassembly_action", ""),
                "evidence": row.get("evidence_ref", ""),
            }
        )

    for row in pass2_rows:
        rows.append(
            {
                "record_type": "replacement_part",
                "id": row.get("entry_id", ""),
                "subject": row.get("item", ""),
                "strategy": row.get("pass2_decision", ""),
                "decision": row.get("sourcing_mode", ""),
                "timing_or_gate": row.get("timing_window", ""),
                "action": row.get("supplier_hint", ""),
                "evidence": row.get("rationale", ""),
            }
        )

    replace_table_sheet(
        workbook,
        "Restore_Replace",
        rows,
        ["record_type", "id", "subject", "strategy", "decision", "timing_or_gate", "action", "evidence"],
    )


def build_procurement_pass2_sheet(workbook, pass2_rows: list[dict[str, str]]) -> None:
    fields = [
        "entry_id",
        "workstream",
        "item",
        "prior_decision",
        "sourcing_mode",
        "timing_window",
        "pass2_decision",
        "budget_mode",
        "basket_id",
        "supplier_hint",
        "rationale",
    ]
    replace_table_sheet(workbook, "Procurement_Pass2", pass2_rows, fields)


def build_tab_reconciliation_sheet(workbook, profile_rows: list[dict[str, str]], backlog_rows: list[dict[str, str]]) -> None:
    issue_count: defaultdict[str, int] = defaultdict(int)
    priority_by_sheet: defaultdict[str, Counter[str]] = defaultdict(Counter)
    for row in backlog_rows:
        sheet = row.get("sheet_name", "")
        issue_count[sheet] += 1
        priority_by_sheet[sheet][row.get("priority", "")] += 1

    profile_map = {row.get("sheet_name", ""): row for row in profile_rows}
    rows: list[dict[str, str]] = []

    for sheet_name in workbook.sheetnames:
        profile = profile_map.get(sheet_name, {})
        priorities = priority_by_sheet.get(sheet_name, Counter())
        top_priority = ""
        for key in ("high", "medium", "low"):
            if priorities.get(key):
                top_priority = key
                break

        state = "active"
        action = "keep"
        if sheet_name in HIDE_REFERENCE_TABS:
            state = "hidden_reference"
            action = "hidden_to_reduce_clutter"
        elif sheet_name in MANAGED_KEEP_TABS:
            state = "managed"
            action = "reconciled_generated_tab"

        rows.append(
            {
                "sheet_name": sheet_name,
                "state": state,
                "issue_count_from_scan": str(issue_count.get(sheet_name, 0)),
                "highest_issue_priority": top_priority,
                "classification": profile.get("classification", ""),
                "structured_rows": profile.get("structured_data_rows", ""),
                "action_taken": action,
            }
        )

    replace_table_sheet(
        workbook,
        "Tab_Reconciliation",
        rows,
        [
            "sheet_name",
            "state",
            "issue_count_from_scan",
            "highest_issue_priority",
            "classification",
            "structured_rows",
            "action_taken",
        ],
    )


def simplify_tabs(workbook) -> tuple[list[str], list[str]]:
    removed: list[str] = []
    hidden: list[str] = []

    for name in MANAGED_DROP_TABS:
        if name in workbook.sheetnames:
            del workbook[name]
            removed.append(name)

    for name in HIDE_REFERENCE_TABS:
        if name in workbook.sheetnames:
            ws = workbook[name]
            ws.sheet_state = "hidden"
            hidden.append(name)

    return removed, hidden


def backup_workbook(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}.full_reconcile_backup_{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def write_report(
    workbook_path: Path,
    backup_path: Path,
    header_notes: list[str],
    removed_tabs: list[str],
    hidden_tabs: list[str],
    status_counts: dict[str, Counter[str]],
) -> None:
    lines: list[str] = []
    lines.append("# Workbook Full Reconciliation")
    lines.append("")
    lines.append(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Workbook updated: `{workbook_path}`")
    lines.append(f"- Backup created: `{backup_path}`")
    lines.append("")
    lines.append("## Header / Structure Fixes")
    lines.append("")
    if not header_notes:
        lines.append("- None")
    else:
        for note in header_notes:
            lines.append(f"- {note}")
    lines.append("")
    lines.append("## Tab Simplification")
    lines.append("")
    lines.append(f"- Removed managed tabs: {', '.join(removed_tabs) if removed_tabs else 'none'}")
    lines.append(f"- Hidden reference tabs: {', '.join(hidden_tabs) if hidden_tabs else 'none'}")
    lines.append("")
    lines.append("## Cost Tab Reconciliation Counts")
    lines.append("")
    for tab, counts in status_counts.items():
        lines.append(f"- `{tab}`")
        for status, count in sorted(counts.items()):
            lines.append(f"  - {status}: {count}")

    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Reconcile all tabs in J40_Costs workbook and simplify managed tab surface.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH, help="Workbook path to update in-place.")
    parser.add_argument("--skip-backup", action="store_true", help="Skip backup creation.")
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    required = [CSV_PROCUREMENT_PASS2, CSV_REASSEMBLY_PACKAGES, CSV_COMPONENT_DISPOSITION, CSV_SHEET_PROFILES, CSV_TIDY_BACKLOG]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required CSV inputs: {', '.join(missing)}")

    backup_path = workbook_path
    if not args.skip_backup:
        backup_path = backup_workbook(workbook_path)

    workbook = load_workbook(workbook_path)
    header_notes = ensure_standard_headers(workbook)
    header_notes.extend(split_parts_mixed_tables(workbook))
    header_notes.extend(move_wiring_suspension_block(workbook))
    header_notes.extend(dedupe_suspension_tab(workbook))
    cost_lines, status_counts = update_cost_tabs_with_reconciled_columns(workbook)

    pass2_rows = load_csv(CSV_PROCUREMENT_PASS2)
    reassembly_rows = load_csv(CSV_REASSEMBLY_PACKAGES)
    component_rows = load_csv(CSV_COMPONENT_DISPOSITION)
    profile_rows = load_csv(CSV_SHEET_PROFILES)
    backlog_rows = load_csv(CSV_TIDY_BACKLOG)

    # Simplify generated tabs before rebuilding the active ones.
    removed_tabs, hidden_tabs = simplify_tabs(workbook)

    build_plan_summary_sheet(workbook, pass2_rows, status_counts)
    build_procurement_pass2_sheet(workbook, pass2_rows)
    build_master_costs_sheet(workbook, cost_lines)
    build_build_plan_sheet(workbook, reassembly_rows, pass2_rows)
    build_restore_replace_sheet(workbook, component_rows, pass2_rows)
    build_tab_reconciliation_sheet(workbook, profile_rows, backlog_rows)

    workbook.save(workbook_path)
    write_report(workbook_path, backup_path, header_notes, removed_tabs, hidden_tabs, status_counts)

    print(f"Updated workbook: {workbook_path}")
    if not args.skip_backup:
        print(f"Backup created: {backup_path}")
    print(f"Reconciled cost lines: {len(cost_lines)}")
    print(f"Removed tabs: {', '.join(removed_tabs) if removed_tabs else 'none'}")
    print(f"Hidden tabs: {', '.join(hidden_tabs) if hidden_tabs else 'none'}")
    print(f"Report: {REPORT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
