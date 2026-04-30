from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")
WASHER_PATTERN = re.compile(r"^m\d+\s*star\s*/\s*lock\s*washer$", re.IGNORECASE)


def normalize(text: object) -> str:
    if text is None:
        return ""
    return " ".join(str(text).strip().split())


def find_or_create_active_plan_insert_row(parts_ws) -> int:
    legacy_row = None
    active_header_row = None
    for r in range(2, parts_ws.max_row + 1):
        item = normalize(parts_ws.cell(r, 1).value)
        if item.lower() == "required purchases (active plan)":
            active_header_row = r
        if item.lower() == "legacy planned items (retained)":
            legacy_row = r
            break
    if legacy_row:
        return legacy_row
    if active_header_row:
        return parts_ws.max_row + 1
    # If active block missing for any reason, append at end.
    return parts_ws.max_row + 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Move wiring fastener rows (nuts/bolts/washers) into Parts tab.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH, help="Workbook path to update in-place.")
    args = parser.parse_args()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = args.workbook.with_name(f"{args.workbook.stem}.move_wiring_fasteners_backup_{ts}{args.workbook.suffix}")
    shutil.copy2(args.workbook, backup)

    wb = load_workbook(args.workbook)
    wiring_ws = wb["Wiring"]
    parts_ws = wb["Parts"]

    candidates: dict[str, dict[str, object]] = {}
    rows_to_delete: list[int] = []

    for r in range(2, wiring_ws.max_row + 1):
        item_raw = wiring_ws.cell(r, 1).value
        item = normalize(item_raw)
        if not item:
            continue
        if not WASHER_PATTERN.match(item):
            continue

        status = normalize(wiring_ws.cell(r, 8).value).lower()
        vendor = wiring_ws.cell(r, 3).value
        price = wiring_ws.cell(r, 2).value
        received = wiring_ws.cell(r, 4).value
        paid = wiring_ws.cell(r, 5).value
        note = wiring_ws.cell(r, 9).value

        # Prefer non-section rows for metadata.
        if item not in candidates or status != "section_header":
            candidates[item] = {
                "price": price,
                "vendor": vendor,
                "received": received,
                "paid": paid,
                "status": status,
                "note": note,
            }

        rows_to_delete.append(r)

    if not candidates:
        wb.save(args.workbook)
        print(f"Updated workbook: {args.workbook}")
        print(f"Backup saved: {backup}")
        print("No washer/fastener rows found in Wiring to move.")
        return

    existing_part_items = {
        normalize(parts_ws.cell(r, 1).value).lower()
        for r in range(2, parts_ws.max_row + 1)
        if normalize(parts_ws.cell(r, 1).value)
    }

    insert_row = find_or_create_active_plan_insert_row(parts_ws)
    added_count = 0

    # Ensure planning column headers exist for readability.
    parts_ws.cell(1, 6).value = "Plan_Timing"
    parts_ws.cell(1, 7).value = "Plan_Decision"

    for item in sorted(candidates.keys()):
        if item.lower() in existing_part_items:
            continue
        meta = candidates[item]
        parts_ws.insert_rows(insert_row, amount=1)
        parts_ws.cell(insert_row, 1).value = item
        parts_ws.cell(insert_row, 2).value = meta["price"]
        parts_ws.cell(insert_row, 3).value = meta["vendor"] or "Montgomery Road (Lahore)"
        parts_ws.cell(insert_row, 4).value = meta["received"]
        parts_ws.cell(insert_row, 5).value = meta["paid"]
        parts_ws.cell(insert_row, 6).value = "pre_order_audit"
        parts_ws.cell(insert_row, 7).value = "stock_audit_then_local_topup"
        parts_ws.cell(insert_row, 11).value = "planned_or_open"
        parts_ws.cell(insert_row, 12).value = (
            "Moved from Wiring tab; fastener item belongs in Parts purchase tracking."
        )
        parts_ws.cell(insert_row, 13).value = None
        insert_row += 1
        added_count += 1

    for r in sorted(set(rows_to_delete), reverse=True):
        wiring_ws.delete_rows(r, 1)

    wb.save(args.workbook)
    print(f"Updated workbook: {args.workbook}")
    print(f"Backup saved: {backup}")
    print(f"Moved fastener rows from Wiring: {len(set(rows_to_delete))}")
    print(f"Added rows to Parts: {added_count}")


if __name__ == "__main__":
    main()
