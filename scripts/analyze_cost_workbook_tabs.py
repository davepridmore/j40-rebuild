from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
MANUAL_DIR = ROOT / "data" / "manual"
DOCS_DIR = ROOT / "docs"

DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")

SHEET_PROFILE_CSV_PATH = MANUAL_DIR / "j40_workbook_sheet_profiles.csv"
TIDY_BACKLOG_CSV_PATH = MANUAL_DIR / "j40_workbook_tidy_backlog.csv"
SHEET_EXPORT_DIR = MANUAL_DIR / "workbook_tabs"
REPORT_MD_PATH = DOCS_DIR / "j40-workbook-tabs-and-tidy-plan.md"

COST_KEYWORDS = {"item", "price", "vendor", "received", "receieved", "paid", "cost"}
ELECTRICAL_KEYWORDS = {
    "wire",
    "loom",
    "connector",
    "pin",
    "circuit",
    "relay",
    "switch",
    "fuse",
    "routing",
    "harness",
}
STATUS_KEYWORDS = {"status", "received", "receieved", "paid", "sent", "used"}
HEADER_HINTS = {
    "type",
    "part",
    "info",
    "reason",
    "category",
    "item",
    "price",
    "vendor",
    "received",
    "receieved",
    "paid",
    "component",
    "quantity",
    "details",
    "section",
    "circuit",
    "detail",
    "colour",
    "color",
    "size",
    "est_length",
    "notes",
    "service",
    "cost",
    "connector",
    "pin",
    "wire_id",
    "function",
    "relay_position",
    "system",
    "fuse",
}
EXPECTED_STATUS_VALUES = {
    "Y",
    "N",
    "?",
    "YES",
    "NO",
    "COD",
    "UNKNOWN",
    "CANCELLED",
    "CANCELLED?",
    "HOLD",
    "DONE",
    "HAVE TO CHECK",
    "HAVETOCHECK",
}
MISSPELLED_HEADERS = {"receieved": "received", "relay assigment": "relay assignment"}


@dataclass(frozen=True)
class SheetProfile:
    sheet_name: str
    classification: str
    max_row: int
    max_column: int
    nonempty_rows: int
    nonempty_cells: int
    header_row: int
    header_count: int
    structured_data_rows: int
    key_headers: str
    recommended_cleanup: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_label(text: str) -> str:
    lowered = text.lower().strip()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    return re.sub(r"_+", "_", lowered).strip("_")


def looks_numeric(text: str) -> bool:
    if not text:
        return False
    normalized = text.replace(",", "").strip()
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", normalized))


def contains_number(text: str) -> bool:
    return bool(re.search(r"\d", text))


def sanitize_filename(name: str) -> str:
    slug = normalize_label(name)
    return slug or "sheet"


def row_values(ws, row_idx: int) -> list[str]:
    return [clean_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]


def header_score(values: list[str], row_idx: int) -> float:
    nonempty = [value for value in values if value]
    if not nonempty:
        return 0.0
    labels = [normalize_label(value) for value in nonempty]
    exact_hint_hits = sum(1 for label in labels if label in HEADER_HINTS)
    non_numeric = sum(1 for value in nonempty if not looks_numeric(value))
    keyword_hits = sum(
        1
        for value in nonempty
        if any(keyword in value.lower() for keyword in COST_KEYWORDS | ELECTRICAL_KEYWORDS | STATUS_KEYWORDS)
    )
    unique = len(set(value.lower() for value in nonempty))
    numeric_ratio = 1.0 - (non_numeric / len(nonempty))
    one_cell_penalty = 2.0 if len(nonempty) == 1 else 0.0
    early_row_bonus = max(0.0, 1.2 - ((row_idx - 1) * 0.12))
    return (
        (exact_hint_hits * 4.0)
        + (keyword_hits * 1.5)
        + (non_numeric * 0.7)
        + (unique * 0.3)
        - (numeric_ratio * 2.0)
        + early_row_bonus
        - one_cell_penalty
    )


def detect_header_row(ws) -> tuple[int, list[str]]:
    best_row = 0
    best_headers: list[str] = []
    best_score = 0.0
    search_limit = min(ws.max_row, 15)
    for row_idx in range(1, search_limit + 1):
        values = row_values(ws, row_idx)
        score = header_score(values, row_idx)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = values

    header_count = len([header for header in best_headers if header])
    if best_row == 0 or header_count < 2 or best_score < 2.5:
        return 0, []
    return best_row, best_headers


def classify_sheet(sheet_name: str, headers: list[str], structured_data_rows: int, max_column: int) -> str:
    sheet_norm = sheet_name.lower()
    header_blob = " ".join(header.lower() for header in headers if header)
    has_cost_shape = "item" in header_blob and ("price" in header_blob or "vendor" in header_blob)
    has_status_shape = any(keyword in header_blob for keyword in STATUS_KEYWORDS)
    is_electrical = any(keyword in sheet_norm for keyword in ELECTRICAL_KEYWORDS) or any(
        keyword in header_blob for keyword in ELECTRICAL_KEYWORDS
    )
    is_mechanical = any(keyword in sheet_norm for keyword in {"suspension", "mechanical"}) or any(
        keyword in header_blob for keyword in {"component", "quantity", "details"}
    )

    if has_cost_shape or sheet_name in {"Tools", "Parts", "Substances", "Wiring", "Service"}:
        return "cost_tracker"
    if is_electrical and structured_data_rows > 0:
        return "electrical_spec"
    if is_mechanical and structured_data_rows > 0:
        return "mechanical_spec"
    if has_status_shape and structured_data_rows > 0:
        return "status_tracking"
    if structured_data_rows == 0 and max_column <= 3:
        return "notes_checklist"
    if max_column <= 3 and structured_data_rows <= 12:
        return "notes_checklist"
    return "mixed_reference"


def recommend_cleanup(classification: str, headers: list[str]) -> str:
    header_blob = " ".join(header.lower() for header in headers if header)
    if classification == "cost_tracker":
        return "Normalize headers and status values, split notes from price fields, then merge into expenses/procurement ledgers."
    if classification == "electrical_spec":
        return "Standardize column headers, isolate narrative rows, and keep spec rows in one normalized table per tab."
    if classification == "mechanical_spec":
        return "Use explicit component/part/quantity/details headers and isolate options vs selected parts."
    if classification == "status_tracking":
        return "Unify sent/received/paid/status tokens and add explicit date+owner columns."
    if classification == "notes_checklist":
        return "Keep as notes but add explicit section headers and remove empty spacer rows."
    if "template" in header_blob:
        return "Remove placeholder rows and convert into explicit template instructions."
    return "Review manually and split mixed content into structured table plus notes."


def export_sheet_csv(ws, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{sanitize_filename(ws.title)}.csv"
    output_path = output_dir / filename

    fieldnames = ["excel_row"] + [f"col_{index}" for index in range(1, ws.max_column + 1)]
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row_idx in range(1, ws.max_row + 1):
            values = row_values(ws, row_idx)
            if not any(values):
                continue
            row: dict[str, str] = {"excel_row": str(row_idx)}
            for col_idx, value in enumerate(values, start=1):
                row[f"col_{col_idx}"] = value
            writer.writerow(row)

    return output_path


def clear_sheet_exports(output_dir: Path) -> None:
    if not output_dir.exists():
        return
    for path in output_dir.glob("*.csv"):
        path.unlink()


def build_tidy_issues(ws, header_row: int, headers: list[str], classification: str, structured_rows: list[list[str]]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    sheet_name = ws.title
    header_labels = [header for header in headers if header]
    normalized_headers = [normalize_label(header) for header in header_labels]

    empty_headers = [str(index) for index, value in enumerate(headers, start=1) if not value]
    if empty_headers and structured_rows:
        issues.append(
            {
                "priority": "high",
                "sheet_name": sheet_name,
                "issue_type": "unnamed_columns",
                "issue_detail": f"Structured data has unnamed columns at positions: {', '.join(empty_headers[:8])}",
                "recommended_action": "Name all active columns and move free-form notes into an explicit notes column.",
            }
        )

    duplicate_headers = [label for label, count in Counter(normalized_headers).items() if label and count > 1]
    if duplicate_headers:
        issues.append(
            {
                "priority": "medium",
                "sheet_name": sheet_name,
                "issue_type": "duplicate_headers",
                "issue_detail": f"Duplicate header labels after normalization: {', '.join(duplicate_headers)}",
                "recommended_action": "Rename duplicate columns to distinct, role-specific names.",
            }
        )

    for header in header_labels:
        normalized = normalize_label(header)
        if normalized in MISSPELLED_HEADERS:
            issues.append(
                {
                    "priority": "medium",
                    "sheet_name": sheet_name,
                    "issue_type": "misspelled_header",
                    "issue_detail": f"Header '{header}' should be '{MISSPELLED_HEADERS[normalized]}'.",
                    "recommended_action": "Correct misspelled headers before normalization.",
                }
            )

    if not structured_rows:
        return issues

    status_col_indexes = [
        index
        for index, header in enumerate(headers, start=1)
        if any(keyword in header.lower() for keyword in STATUS_KEYWORDS)
    ]
    item_col = next((index for index, header in enumerate(headers, start=1) if "item" in header.lower() or "service" in header.lower()), None)
    price_col = next((index for index, header in enumerate(headers, start=1) if "price" in header.lower() or "cost" in header.lower()), None)
    vendor_col = next((index for index, header in enumerate(headers, start=1) if "vendor" in header.lower()), None)
    received_col = next((index for index, header in enumerate(headers, start=1) if "recei" in header.lower()), None)
    paid_col = next((index for index, header in enumerate(headers, start=1) if "paid" in header.lower()), None)

    schema_mismatch_rows = 0
    cost_shape_rows: list[list[str]] = []
    if classification == "cost_tracker" and item_col and price_col and vendor_col:
        for row in structured_rows:
            item_value = row[item_col - 1].strip() if item_col <= len(row) else ""
            price_value = row[price_col - 1].strip() if price_col <= len(row) else ""
            vendor_value = row[vendor_col - 1].strip() if vendor_col <= len(row) else ""
            received_value = row[received_col - 1].strip() if received_col and received_col <= len(row) else ""
            paid_value = row[paid_col - 1].strip() if paid_col and paid_col <= len(row) else ""

            long_status_text = any(
                len(value) > 14 and re.sub(r"\s+", " ", value.upper()) not in EXPECTED_STATUS_VALUES
                for value in [received_value, paid_value]
                if value
            )
            part_code_in_price = bool(re.search(r"[A-Za-z]", price_value)) and bool(re.search(r"\d", price_value))
            qty_like_vendor = looks_numeric(vendor_value)
            likely_schema_mismatch = bool(item_value) and ((part_code_in_price and qty_like_vendor) or long_status_text)

            if likely_schema_mismatch:
                schema_mismatch_rows += 1
            else:
                cost_shape_rows.append(row)

        if schema_mismatch_rows > 0:
            issues.append(
                {
                    "priority": "high",
                    "sheet_name": sheet_name,
                    "issue_type": "mixed_schema_rows",
                    "issue_detail": f"{schema_mismatch_rows} rows do not fit the Item/Price/Vendor/Received/Paid schema.",
                    "recommended_action": "Move those rows to the correct sheet (for example suspension/spec tabs) and keep this tab cost-only.",
                }
            )
    else:
        cost_shape_rows = structured_rows

    unexpected_tokens: set[str] = set()
    for row in cost_shape_rows:
        for col_idx in status_col_indexes:
            value = row[col_idx - 1].strip()
            if not value:
                continue
            token = re.sub(r"\s+", " ", value.upper())
            if token not in EXPECTED_STATUS_VALUES:
                unexpected_tokens.add(token)
    if unexpected_tokens:
        issues.append(
            {
                "priority": "high",
                "sheet_name": sheet_name,
                "issue_type": "inconsistent_status_tokens",
                "issue_detail": f"Unexpected status tokens: {', '.join(sorted(unexpected_tokens)[:12])}",
                "recommended_action": "Map status values to canonical tokens (Y, N, ?, COD, CANCELLED, HAVE TO CHECK).",
            }
        )

    price_col_indexes = [
        index
        for index, header in enumerate(headers, start=1)
        if any(keyword in header.lower() for keyword in {"price", "cost", "amount"})
    ]
    mixed_price_cells = 0
    non_numeric_price_cells = 0
    for row in cost_shape_rows:
        for col_idx in price_col_indexes:
            value = row[col_idx - 1].strip()
            if not value:
                continue
            if contains_number(value) and not looks_numeric(value):
                mixed_price_cells += 1
            elif not contains_number(value):
                non_numeric_price_cells += 1
    if mixed_price_cells > 0:
        issues.append(
            {
                "priority": "medium",
                "sheet_name": sheet_name,
                "issue_type": "mixed_price_cells",
                "issue_detail": f"{mixed_price_cells} price cells contain numeric values mixed with text.",
                "recommended_action": "Split numeric amount into amount column and move text into notes/status columns.",
            }
        )
    if non_numeric_price_cells > 0:
        issues.append(
            {
                "priority": "low",
                "sheet_name": sheet_name,
                "issue_type": "non_numeric_price_cells",
                "issue_detail": f"{non_numeric_price_cells} price cells contain no numeric value.",
                "recommended_action": "Fill missing amounts or mark as amount_status=missing.",
            }
        )

    if classification == "cost_tracker":
        if item_col:
            item_prices: defaultdict[str, set[str]] = defaultdict(set)
            for row in cost_shape_rows:
                item = row[item_col - 1].strip()
                if not item:
                    continue
                if price_col:
                    price = row[price_col - 1].strip()
                    if price:
                        item_prices[normalize_label(item)].add(price)
            variable_priced_duplicates = [item for item, prices in item_prices.items() if len(prices) > 1]
            if variable_priced_duplicates:
                issues.append(
                    {
                        "priority": "medium",
                        "sheet_name": sheet_name,
                        "issue_type": "duplicate_item_with_different_prices",
                        "issue_detail": f"{len(variable_priced_duplicates)} items appear with multiple prices.",
                        "recommended_action": "Keep one active row per item and move alternatives/history into quotes or notes.",
                    }
                )

    return issues


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(workbook_path: Path, profiles: list[SheetProfile], issues: list[dict[str, str]]) -> None:
    class_counts = Counter(profile.classification for profile in profiles)
    issue_counts = Counter(issue["priority"] for issue in issues)
    issues_by_sheet = Counter(issue["sheet_name"] for issue in issues)

    lines: list[str] = []
    lines.append("# J40 Workbook Tabs: Understanding + Tidy Plan")
    lines.append("")
    lines.append(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Workbook: `{workbook_path}`")
    lines.append(f"- Sheet count: {len(profiles)}")
    lines.append("- Sheet profile CSV: `data/manual/j40_workbook_sheet_profiles.csv`")
    lines.append("- Tidy backlog CSV: `data/manual/j40_workbook_tidy_backlog.csv`")
    lines.append("- Per-sheet exports: `data/manual/workbook_tabs/*.csv`")
    lines.append("")
    lines.append("## Sheet Types")
    lines.append("")
    for category, count in sorted(class_counts.items()):
        lines.append(f"- `{category}`: {count}")
    lines.append("")
    lines.append("## Tidy Backlog Summary")
    lines.append("")
    lines.append(f"- `high`: {issue_counts.get('high', 0)}")
    lines.append(f"- `medium`: {issue_counts.get('medium', 0)}")
    lines.append(f"- `low`: {issue_counts.get('low', 0)}")
    lines.append("")
    lines.append("## Sheets With Most Tidy Work")
    lines.append("")
    if issues_by_sheet:
        for sheet_name, count in issues_by_sheet.most_common(10):
            lines.append(f"- `{sheet_name}`: {count} issues")
    else:
        lines.append("- No issues detected by automated checks.")
    lines.append("")
    lines.append("## Sheet Inventory")
    lines.append("")
    for profile in profiles:
        lines.append(
            f"- `{profile.sheet_name}` [{profile.classification}] "
            f"rows={profile.max_row}, cols={profile.max_column}, nonempty_rows={profile.nonempty_rows}, "
            f"structured_rows={profile.structured_data_rows}"
        )

    REPORT_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze all tabs in J40_Costs.xlsx and build a tidy backlog.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to J40 costs workbook (.xlsx).",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    clear_sheet_exports(SHEET_EXPORT_DIR)

    profiles: list[SheetProfile] = []
    issues: list[dict[str, str]] = []

    for ws in workbook.worksheets:
        export_sheet_csv(ws, SHEET_EXPORT_DIR)

        nonempty_rows = 0
        nonempty_cells = 0
        all_rows: list[list[str]] = []
        for row_idx in range(1, ws.max_row + 1):
            values = row_values(ws, row_idx)
            all_rows.append(values)
            if any(values):
                nonempty_rows += 1
                nonempty_cells += sum(1 for value in values if value)

        header_row, headers = detect_header_row(ws)
        structured_rows: list[list[str]] = []
        if header_row:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                values = all_rows[row_idx - 1]
                if any(values):
                    structured_rows.append(values)

        classification = classify_sheet(ws.title, headers, len(structured_rows), ws.max_column)
        cleanup_note = recommend_cleanup(classification, headers)
        key_headers = " | ".join([header for header in headers if header][:8])

        profiles.append(
            SheetProfile(
                sheet_name=ws.title,
                classification=classification,
                max_row=ws.max_row,
                max_column=ws.max_column,
                nonempty_rows=nonempty_rows,
                nonempty_cells=nonempty_cells,
                header_row=header_row,
                header_count=len([header for header in headers if header]),
                structured_data_rows=len(structured_rows),
                key_headers=key_headers,
                recommended_cleanup=cleanup_note,
            )
        )

        issues.extend(build_tidy_issues(ws, header_row, headers, classification, structured_rows))

    write_csv(
        SHEET_PROFILE_CSV_PATH,
        [profile.__dict__ for profile in profiles],
        [
            "sheet_name",
            "classification",
            "max_row",
            "max_column",
            "nonempty_rows",
            "nonempty_cells",
            "header_row",
            "header_count",
            "structured_data_rows",
            "key_headers",
            "recommended_cleanup",
        ],
    )

    write_csv(
        TIDY_BACKLOG_CSV_PATH,
        issues,
        ["priority", "sheet_name", "issue_type", "issue_detail", "recommended_action"],
    )

    write_report(workbook_path, profiles, issues)

    print(f"Workbook: {workbook_path}")
    print(f"Wrote sheet profiles: {SHEET_PROFILE_CSV_PATH.relative_to(ROOT)}")
    print(f"Wrote tidy backlog: {TIDY_BACKLOG_CSV_PATH.relative_to(ROOT)}")
    print(f"Wrote report: {REPORT_MD_PATH.relative_to(ROOT)}")
    print(f"Exported sheet CSVs: {SHEET_EXPORT_DIR.relative_to(ROOT)}")
    print(f"Sheets analyzed: {len(profiles)}")
    print(f"Tidy issues found: {len(issues)}")


if __name__ == "__main__":
    main()
