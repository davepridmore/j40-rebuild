from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
MANUAL_DIR = ROOT / "data" / "manual"
DOCS_DIR = ROOT / "docs"

DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")

TIDY_CSV_PATH = MANUAL_DIR / "j40_costs_cost_tabs_tidy.csv"
TIDY_REPORT_PATH = DOCS_DIR / "j40-costs-tidy-extract.md"

COST_SHEETS = ("Tools", "Parts", "Substances", "Wiring", "Service")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def parse_amount(value: str) -> tuple[str, str]:
    if not value:
        return "", ""
    normalized = value.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", normalized)
    if not match:
        return "", value
    number = match.group(0)
    amount = str(int(float(number))) if "." in number else number
    note = normalized.replace(number, "").strip(" -")
    return amount, note


def normalize_status(value: str) -> str:
    token = re.sub(r"\s+", " ", value.upper().strip())
    if not token:
        return ""
    if token in {"Y", "YES", "PAID"}:
        return "yes"
    if token in {"N", "NO"}:
        return "no"
    if token == "?":
        return "unknown"
    if "COD" in token:
        return "cod"
    if "CANCEL" in token:
        return "cancelled"
    if "CHECK" in token:
        return "needs_check"
    if token in {"UNKNOWN", "HOLD"}:
        return "unknown"
    return "other"


def looks_numeric(value: str) -> bool:
    if not value:
        return False
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", value.replace(",", "").strip()))


def detect_header_indexes(ws) -> tuple[int, dict[str, int]]:
    headers = [clean_text(ws.cell(1, index).value).lower() for index in range(1, ws.max_column + 1)]
    indexes: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        if header in {"item", "service", "part"} and "item" not in indexes:
            indexes["item"] = index
        if header in {"price", "cost"} and "price" not in indexes:
            indexes["price"] = index
        if header in {"vendor", "company"} and "vendor" not in indexes:
            indexes["vendor"] = index
        if header in {"received", "receieved"} and "received" not in indexes:
            indexes["received"] = index
        if header == "paid" and "paid" not in indexes:
            indexes["paid"] = index
    return 1, indexes


def row_disposition(item: str, price: str, vendor: str, received: str, paid: str) -> tuple[str, str]:
    if item and not any([price, vendor, received, paid]):
        return "section_header", ""

    long_status_text = any(
        len(value) > 14 and normalize_status(value) in {"other", ""}
        for value in [received, paid]
        if value
    )
    part_code_in_price = bool(re.search(r"[A-Za-z]", price)) and bool(re.search(r"\d", price))
    qty_like_vendor = looks_numeric(vendor)
    if item and ((part_code_in_price and qty_like_vendor) or long_status_text):
        target = ""
        joined = f"{item} {received} {paid}".lower()
        if any(keyword in joined for keyword in {"spring", "shock", "shackle", "suspension", "leaf"}):
            target = "Suspension"
        return "schema_mismatch", target

    return "line_item", ""


def tidy_workbook(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    tidy_rows: list[dict[str, str]] = []

    for sheet_name in COST_SHEETS:
        ws = workbook[sheet_name]
        header_row, indexes = detect_header_indexes(ws)
        item_col = indexes.get("item", 1)
        price_col = indexes.get("price", 2)
        vendor_col = indexes.get("vendor", 3)
        received_col = indexes.get("received", 4)
        paid_col = indexes.get("paid", 5)

        for row_index in range(header_row + 1, ws.max_row + 1):
            item = clean_text(ws.cell(row_index, item_col).value)
            price_raw = clean_text(ws.cell(row_index, price_col).value)
            vendor = clean_text(ws.cell(row_index, vendor_col).value)
            received_raw = clean_text(ws.cell(row_index, received_col).value)
            paid_raw = clean_text(ws.cell(row_index, paid_col).value)

            extra_notes: list[str] = []
            for extra_col in range(max(item_col, price_col, vendor_col, received_col, paid_col) + 1, ws.max_column + 1):
                note = clean_text(ws.cell(row_index, extra_col).value)
                if note:
                    extra_notes.append(note)

            if not any([item, price_raw, vendor, received_raw, paid_raw, extra_notes]):
                continue

            amount, amount_note = parse_amount(price_raw)
            disposition, suggested_target = row_disposition(item, price_raw, vendor, received_raw, paid_raw)

            tidy_rows.append(
                {
                    "source_sheet": sheet_name,
                    "source_row": str(row_index),
                    "item": item,
                    "price_raw": price_raw,
                    "amount": amount,
                    "amount_note": amount_note,
                    "vendor": vendor,
                    "received_raw": received_raw,
                    "received_status": normalize_status(received_raw),
                    "paid_raw": paid_raw,
                    "paid_status": normalize_status(paid_raw),
                    "extra_notes": " | ".join(extra_notes),
                    "row_disposition": disposition,
                    "suggested_target_tab": suggested_target,
                }
            )

    return tidy_rows


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(workbook_path: Path, rows: list[dict[str, str]]) -> None:
    by_sheet = Counter(row["source_sheet"] for row in rows)
    by_disposition = Counter(row["row_disposition"] for row in rows)
    by_received = Counter(row["received_status"] for row in rows if row["row_disposition"] == "line_item")
    by_paid = Counter(row["paid_status"] for row in rows if row["row_disposition"] == "line_item")
    schema_mismatch_rows = [row for row in rows if row["row_disposition"] == "schema_mismatch"]

    lines: list[str] = []
    lines.append("# J40 Costs Workbook Tidy Extract")
    lines.append("")
    lines.append(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Workbook: `{workbook_path}`")
    lines.append("- Output: `data/manual/j40_costs_cost_tabs_tidy.csv`")
    lines.append(f"- Total extracted rows: {len(rows)}")
    lines.append("")
    lines.append("## Rows By Sheet")
    lines.append("")
    for sheet, count in sorted(by_sheet.items()):
        lines.append(f"- `{sheet}`: {count}")
    lines.append("")
    lines.append("## Row Disposition")
    lines.append("")
    for disposition, count in sorted(by_disposition.items()):
        lines.append(f"- `{disposition}`: {count}")
    lines.append("")
    lines.append("## Normalized Received Status (line-item only)")
    lines.append("")
    for status, count in sorted(by_received.items()):
        lines.append(f"- `{status or 'blank'}`: {count}")
    lines.append("")
    lines.append("## Normalized Paid Status (line-item only)")
    lines.append("")
    for status, count in sorted(by_paid.items()):
        lines.append(f"- `{status or 'blank'}`: {count}")
    lines.append("")
    lines.append("## Schema-Mismatch Rows (likely wrong tab)")
    lines.append("")
    if not schema_mismatch_rows:
        lines.append("- None")
    else:
        for row in schema_mismatch_rows[:25]:
            lines.append(
                f"- `{row['source_sheet']}#{row['source_row']}` {row['item']} "
                f"(suggested target: {row['suggested_target_tab'] or 'review'})"
            )

    TIDY_REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a tidy normalized extract from J40 cost sheets.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to J40 costs workbook (.xlsx).",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    rows = tidy_workbook(workbook_path)
    write_csv(
        TIDY_CSV_PATH,
        rows,
        [
            "source_sheet",
            "source_row",
            "item",
            "price_raw",
            "amount",
            "amount_note",
            "vendor",
            "received_raw",
            "received_status",
            "paid_raw",
            "paid_status",
            "extra_notes",
            "row_disposition",
            "suggested_target_tab",
        ],
    )
    write_report(workbook_path, rows)

    print(f"Workbook: {workbook_path}")
    print(f"Wrote tidy extract: {TIDY_CSV_PATH.relative_to(ROOT)}")
    print(f"Wrote tidy report: {TIDY_REPORT_PATH.relative_to(ROOT)}")
    print(f"Rows extracted: {len(rows)}")


if __name__ == "__main__":
    main()
