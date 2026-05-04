from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
MANUAL_DIR = ROOT / "data" / "manual"
DOCS_DIR = ROOT / "docs"

DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")

WORKBOOK_ROWS_CSV_PATH = MANUAL_DIR / "j40_costs_workbook_rows.csv"
RECON_CSV_PATH = MANUAL_DIR / "j40_costs_expenses_reconciliation.csv"
REPORT_MD_PATH = DOCS_DIR / "j40-costs-workbook-reconciliation.md"
EXPENSES_PATH = MANUAL_DIR / "expenses.csv"

KNOWN_COST_SHEETS = {"Tools", "Parts", "Substances", "Wiring", "Service"}
HEADER_ALIASES = {
    "item": {"item", "service", "part"},
    "price": {"price", "cost"},
    "vendor": {"vendor", "company"},
    "received": {"received", "receieved"},
    "paid": {"paid"},
}


@dataclass(frozen=True)
class WorkbookRow:
    source_sheet: str
    source_row: int
    item: str
    price_raw: str
    parsed_amount: str
    vendor: str
    received_flag: str
    paid_flag: str
    notes: str
    row_type: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def norm_text(value: str) -> str:
    lowered = value.lower()
    lowered = lowered.replace("&", " and ")
    lowered = re.sub(r"[^a-z0-9]+", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def is_heat_glow_plug_item(value: str) -> bool:
    normalized = norm_text(value)
    tokens = set(normalized.split())
    has_plug = bool(tokens & {"plug", "plugs"})
    return has_plug and bool(tokens & {"heat", "glow"})


def workbook_supply_ref(source_sheet: str, source_row: int) -> str:
    sheet_token = norm_text(source_sheet)
    source_name_by_sheet = {
        "tools": "workbook_tools",
        "parts": "workbook_parts",
        "substances": "workbook_substances",
    }
    source_name = source_name_by_sheet.get(sheet_token)
    return f"{source_name}#row_{source_row}" if source_name else ""


def evidence_contains_source_ref(evidence_ref: str, source_ref: str) -> bool:
    if not evidence_ref or not source_ref:
        return False
    tokens = [token.strip() for token in re.split(r"[|,;]", evidence_ref)]
    return source_ref in tokens


def apply_match(result: dict[str, str], status: str, score: str, chosen: dict[str, str]) -> dict[str, str]:
    result.update(
        {
            "match_status": status,
            "match_score": score,
            "matched_entry_id": chosen.get("entry_id", ""),
            "matched_item": chosen.get("item", ""),
            "matched_company": chosen.get("company", ""),
            "matched_bucket": chosen.get("bucket", ""),
            "matched_phase": chosen.get("phase", ""),
            "matched_workstream": chosen.get("workstream", ""),
            "matched_procurement_stage": chosen.get("procurement_stage", ""),
        }
    )
    return result


def parse_amount(value: str) -> str:
    if not value:
        return ""
    normalized = value.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", normalized)
    if not match:
        return ""
    number = match.group(0)
    if "." in number:
        return str(int(float(number)))
    return number


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_map: dict[str, int] = {}
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        values = [clean_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        lowered = [value.lower() for value in values]
        local: dict[str, int] = {}
        for col_idx, value in enumerate(lowered, start=1):
            for key, aliases in HEADER_ALIASES.items():
                if value in aliases and key not in local:
                    local[key] = col_idx
        if len(local) > len(best_map):
            best_map = local
            best_row = row_idx
    return best_row, best_map


def extract_sheet_rows(ws) -> list[WorkbookRow]:
    header_row, header_map = find_header_row(ws)
    is_known_cost_sheet = ws.title in KNOWN_COST_SHEETS
    has_cost_shape = "item" in header_map and ("price" in header_map or "vendor" in header_map)
    if not (is_known_cost_sheet or has_cost_shape):
        return []

    item_col = header_map.get("item", 1)
    price_col = header_map.get("price", 2)
    vendor_col = header_map.get("vendor", 3)
    received_col = header_map.get("received", 4)
    paid_col = header_map.get("paid", 5)
    start_row = header_row + 1 if has_cost_shape else 2

    extracted: list[WorkbookRow] = []
    for row_idx in range(start_row, ws.max_row + 1):
        item = clean_text(ws.cell(row_idx, item_col).value)
        price_raw = clean_text(ws.cell(row_idx, price_col).value)
        vendor = clean_text(ws.cell(row_idx, vendor_col).value)
        received = clean_text(ws.cell(row_idx, received_col).value)
        paid = clean_text(ws.cell(row_idx, paid_col).value)

        trailing_notes: list[str] = []
        for col_idx in range(max(item_col, price_col, vendor_col, received_col, paid_col) + 1, ws.max_column + 1):
            note_piece = clean_text(ws.cell(row_idx, col_idx).value)
            if note_piece:
                trailing_notes.append(note_piece)
        notes = " | ".join(trailing_notes)

        if not any([item, price_raw, vendor, received, paid, notes]):
            continue

        row_type = "line_item"
        if item and not any([price_raw, vendor, received, paid]):
            row_type = "section_header"

        extracted.append(
            WorkbookRow(
                source_sheet=ws.title,
                source_row=row_idx,
                item=item,
                price_raw=price_raw,
                parsed_amount=parse_amount(price_raw),
                vendor=vendor,
                received_flag=received,
                paid_flag=paid,
                notes=notes,
                row_type=row_type,
            )
        )

    return extracted


def load_expenses() -> list[dict[str, str]]:
    with EXPENSES_PATH.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def vendor_match(workbook_vendor: str, expense_vendor: str) -> bool:
    left = norm_text(workbook_vendor)
    right = norm_text(expense_vendor)
    if not left or not right:
        return False
    if left == right:
        return True
    if left in right or right in left:
        return True
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    shared = left_tokens & right_tokens
    return len(shared) >= 2


def find_best_match(workbook_row: WorkbookRow, expenses_rows: list[dict[str, str]]) -> dict[str, str]:
    result = {
        "match_status": "no_match",
        "match_score": "",
        "matched_entry_id": "",
        "matched_item": "",
        "matched_company": "",
        "matched_bucket": "",
        "matched_phase": "",
        "matched_workstream": "",
        "matched_procurement_stage": "",
    }

    item_norm = norm_text(workbook_row.item)
    if not item_norm or workbook_row.row_type != "line_item":
        result["match_status"] = "not_applicable"
        return result

    exact_candidates = [row for row in expenses_rows if norm_text(row.get("item", "")) == item_norm]
    if exact_candidates:
        vendor_exact = [row for row in exact_candidates if vendor_match(workbook_row.vendor, row.get("company", ""))]
        chosen = vendor_exact[0] if vendor_exact else exact_candidates[0]
        return apply_match(result, "matched_exact_item_vendor" if vendor_exact else "matched_exact_item", "1.0", chosen)

    explicit_ref = workbook_supply_ref(workbook_row.source_sheet, workbook_row.source_row)
    if explicit_ref:
        evidence_candidates = [
            row
            for row in expenses_rows
            if evidence_contains_source_ref(row.get("evidence_ref", ""), explicit_ref)
        ]
        if len(evidence_candidates) == 1:
            return apply_match(result, "matched_manual_evidence_ref", "1.0", evidence_candidates[0])

    note_norm = norm_text(f"{workbook_row.item} {workbook_row.notes}")
    if "total bi metal hole saw 22mm tac410221" in note_norm or "22mm bi metal hole saw" in note_norm:
        hole_saw_candidates = [
            row for row in expenses_rows if row.get("entry_id") == "tool_total_bi_metal_hole_saw_22mm_tac410221"
        ]
        if hole_saw_candidates:
            return apply_match(result, "matched_fulfilled_by_toolsmart_item", "1.0", hole_saw_candidates[0])

    if "total round steel file 200mm tht91386" in note_norm:
        file_candidates = [
            row for row in expenses_rows if row.get("entry_id") == "tool_total_round_steel_file_200mm_tht91386"
        ]
        if file_candidates:
            return apply_match(result, "matched_fulfilled_by_toolsmart_item", "1.0", file_candidates[0])

    if is_heat_glow_plug_item(workbook_row.item):
        alias_candidates = [
            row
            for row in expenses_rows
            if row.get("bucket", "").strip().lower() == "parts" and is_heat_glow_plug_item(row.get("item", ""))
        ]
        if len(alias_candidates) == 1:
            return apply_match(result, "matched_alias_heat_glow_plugs", "1.0", alias_candidates[0])

    best_row: dict[str, str] | None = None
    best_score = 0.0
    for expense in expenses_rows:
        candidate_item = expense.get("item", "")
        candidate_norm = norm_text(candidate_item)
        if not candidate_norm:
            continue
        score = SequenceMatcher(None, item_norm, candidate_norm).ratio()
        if score > best_score:
            best_score = score
            best_row = expense

    if best_row and best_score >= 0.86:
        return apply_match(result, "matched_fuzzy", f"{best_score:.3f}", best_row)

    return result


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(
    workbook_path: Path,
    workbook_rows: list[WorkbookRow],
    reconciliation_rows: list[dict[str, str]],
) -> None:
    rows_by_sheet = Counter(row.source_sheet for row in workbook_rows)
    row_type_counts = Counter(row.row_type for row in workbook_rows)
    match_counts = Counter(row["match_status"] for row in reconciliation_rows)

    unmatched_priority = [
        row
        for row in reconciliation_rows
        if row["match_status"] == "no_match" and (row["parsed_amount"] or row["received_flag"].upper() == "Y" or row["paid_flag"].upper() == "Y")
    ]

    lines: list[str] = []
    lines.append("# J40 Costs Workbook Reconciliation")
    lines.append("")
    lines.append(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Workbook: `{workbook_path}`")
    lines.append(f"- Extracted rows: {len(workbook_rows)}")
    lines.append(f"- Line-item rows: {row_type_counts.get('line_item', 0)}")
    lines.append(f"- Section-header rows: {row_type_counts.get('section_header', 0)}")
    lines.append("- Workbook rows CSV: `data/manual/j40_costs_workbook_rows.csv`")
    lines.append("- Reconciliation CSV: `data/manual/j40_costs_expenses_reconciliation.csv`")
    lines.append("")
    lines.append("## Extracted Rows By Sheet")
    lines.append("")
    for sheet_name, count in rows_by_sheet.most_common():
        lines.append(f"- `{sheet_name}`: {count}")
    lines.append("")
    lines.append("## Match Status")
    lines.append("")
    for status, count in sorted(match_counts.items()):
        lines.append(f"- `{status}`: {count}")
    lines.append("")
    lines.append("## Priority Workbook Rows Still Unmatched")
    lines.append("")
    if not unmatched_priority:
        lines.append("- None")
    else:
        for row in unmatched_priority[:25]:
            lines.append(
                f"- `{row['source_sheet']}#{row['source_row']}` {row['item']} "
                f"(price={row['parsed_amount'] or 'missing'}, vendor={row['vendor'] or 'n/a'})"
            )

    REPORT_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Register and reconcile J40 cost workbook data against expenses.csv")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to J40 costs workbook (.xlsx).",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    workbook_rows: list[WorkbookRow] = []
    for ws in workbook.worksheets:
        workbook_rows.extend(extract_sheet_rows(ws))

    workbook_rows_csv = [
        {
            "source_workbook_path": str(workbook_path),
            "source_sheet": row.source_sheet,
            "source_row": str(row.source_row),
            "item": row.item,
            "price_raw": row.price_raw,
            "parsed_amount": row.parsed_amount,
            "currency": "PKR",
            "vendor": row.vendor,
            "received_flag": row.received_flag,
            "paid_flag": row.paid_flag,
            "notes": row.notes,
            "row_type": row.row_type,
        }
        for row in workbook_rows
    ]
    write_csv(
        WORKBOOK_ROWS_CSV_PATH,
        workbook_rows_csv,
        [
            "source_workbook_path",
            "source_sheet",
            "source_row",
            "item",
            "price_raw",
            "parsed_amount",
            "currency",
            "vendor",
            "received_flag",
            "paid_flag",
            "notes",
            "row_type",
        ],
    )

    expenses_rows = load_expenses()
    recon_rows: list[dict[str, str]] = []
    for row in workbook_rows:
        match = find_best_match(row, expenses_rows)
        recon_rows.append(
            {
                "source_sheet": row.source_sheet,
                "source_row": str(row.source_row),
                "item": row.item,
                "parsed_amount": row.parsed_amount,
                "vendor": row.vendor,
                "received_flag": row.received_flag,
                "paid_flag": row.paid_flag,
                "workbook_row_type": row.row_type,
                "match_status": match["match_status"],
                "match_score": match["match_score"],
                "matched_entry_id": match["matched_entry_id"],
                "matched_item": match["matched_item"],
                "matched_company": match["matched_company"],
                "matched_bucket": match["matched_bucket"],
                "matched_phase": match["matched_phase"],
                "matched_workstream": match["matched_workstream"],
                "matched_procurement_stage": match["matched_procurement_stage"],
            }
        )

    write_csv(
        RECON_CSV_PATH,
        recon_rows,
        [
            "source_sheet",
            "source_row",
            "item",
            "parsed_amount",
            "vendor",
            "received_flag",
            "paid_flag",
            "workbook_row_type",
            "match_status",
            "match_score",
            "matched_entry_id",
            "matched_item",
            "matched_company",
            "matched_bucket",
            "matched_phase",
            "matched_workstream",
            "matched_procurement_stage",
        ],
    )

    write_report(workbook_path, workbook_rows, recon_rows)

    print(f"Workbook: {workbook_path}")
    print(f"Wrote workbook rows: {WORKBOOK_ROWS_CSV_PATH.relative_to(ROOT)}")
    print(f"Wrote reconciliation: {RECON_CSV_PATH.relative_to(ROOT)}")
    print(f"Wrote report: {REPORT_MD_PATH.relative_to(ROOT)}")
    print(f"Rows extracted: {len(workbook_rows)}")


if __name__ == "__main__":
    main()
