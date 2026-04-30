from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
MANUAL_DIR = ROOT / "data" / "manual"
EXPENSES_PATH = MANUAL_DIR / "expenses.csv"
OUTPUT_PATH = MANUAL_DIR / "j40-master-tracker.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="0F243E")

SHEET_NOTES = {
    "Initial Price": "Known quote/initial price points (PKR > 0, confirmed amount rows).",
    "Purchase Registration": "Operational register for purchase state reconciliation.",
    "Purchase of Goods": "Rows classified as goods (tools + parts).",
    "Purchase of Services": "Rows classified as services (labour + admin).",
}


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def parse_decimal(raw: str) -> Decimal | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def autosize_sheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 70)


def style_header_row(worksheet, row_number: int) -> None:
    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def company_value(row: dict[str, str]) -> str:
    value = (row.get("company") or row.get("vendor") or "").strip()
    return value if value else "TBD (confirm)"


def classify_purchase_type(row: dict[str, str]) -> str:
    bucket = (row.get("bucket") or "").strip().lower()
    if bucket in {"tools", "parts"}:
        return "goods"
    if bucket in {"labour", "admin"}:
        return "services"
    return "services"


def derive_purchase_registration(row: dict[str, str]) -> str:
    status = (row.get("status") or "").strip().lower()
    payment_status = (row.get("payment_status") or "").strip().lower()
    delivery_status = (row.get("delivery_status") or "").strip().lower()
    procurement_stage = (row.get("procurement_stage") or "").strip().lower()

    if delivery_status in {"received", "installed", "completed"} or status in {"paid", "installed", "received", "credited"}:
        if payment_status in {"unknown", "not_paid", ""} or delivery_status == "needs_confirmation" or procurement_stage == "received_candidate":
            return "registered_needs_confirmation"
        return "registered"
    if delivery_status == "needs_confirmation" or procurement_stage == "received_candidate":
        return "registered_needs_confirmation"
    if status == "quote":
        return "quote_only"
    if status == "planned":
        return "planned"
    if status == "researching" or delivery_status == "researching" or procurement_stage == "researching":
        return "researching"
    return "open"


def write_rows_sheet(workbook: Workbook, title: str, header: list[str], rows: list[list[str]]) -> int:
    worksheet = workbook.create_sheet(title)
    worksheet.append(header)
    style_header_row(worksheet, 1)

    numeric_columns: set[int] = set()
    for index, column_name in enumerate(header, start=1):
        normalized = column_name.strip().lower()
        if normalized == "amount" or normalized.endswith("_pkr"):
            numeric_columns.add(index)

    for row in rows:
        worksheet.append(row)
        for column_index in numeric_columns:
            cell = worksheet.cell(row=worksheet.max_row, column=column_index)
            value = parse_decimal(str(cell.value or ""))
            if value is not None:
                cell.value = float(value)
                cell.number_format = '#,##0.00'

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_sheet(worksheet)
    return len(rows)


def build_initial_price_rows(expenses: list[dict[str, str]]) -> tuple[list[str], list[list[str]]]:
    header = [
        "entry_id",
        "item",
        "purchase_type",
        "initial_price_pkr",
        "currency",
        "price_source",
        "company_name",
        "payment_status",
        "delivery_status",
        "procurement_stage",
        "evidence_ref",
    ]

    rows: list[list[str]] = []
    for row in expenses:
        amount = parse_decimal(row.get("amount", ""))
        if amount is None or amount <= 0:
            continue
        if (row.get("amount_status") or "").strip().lower() != "confirmed":
            continue

        status = (row.get("status") or "").strip().lower()
        if status == "quote":
            price_source = "quoted_initial"
        elif status in {"paid", "received", "installed", "credited"}:
            price_source = "registered_purchase"
        else:
            price_source = status or "captured"

        rows.append(
            [
                row.get("entry_id", ""),
                row.get("item", ""),
                classify_purchase_type(row),
                row.get("amount", ""),
                row.get("currency", ""),
                price_source,
                company_value(row),
                row.get("payment_status", ""),
                row.get("delivery_status", ""),
                row.get("procurement_stage", ""),
                row.get("evidence_ref", ""),
            ]
        )

    rows.sort(key=lambda r: (r[2], r[1], r[0]))
    return header, rows


def build_purchase_registration_rows(expenses: list[dict[str, str]]) -> tuple[list[str], list[list[str]]]:
    header = [
        "entry_id",
        "date",
        "item",
        "purchase_type",
        "purchase_registration",
        "payment_status",
        "delivery_status",
        "procurement_stage",
        "amount",
        "currency",
        "company_name",
        "evidence_ref",
        "notes",
    ]

    priority = {
        "registered": 0,
        "registered_needs_confirmation": 1,
        "quote_only": 2,
        "planned": 3,
        "researching": 4,
        "open": 5,
    }

    rows: list[list[str]] = []
    for row in expenses:
        registration = derive_purchase_registration(row)
        rows.append(
            [
                row.get("entry_id", ""),
                row.get("date", ""),
                row.get("item", ""),
                classify_purchase_type(row),
                registration,
                row.get("payment_status", ""),
                row.get("delivery_status", ""),
                row.get("procurement_stage", ""),
                row.get("amount", ""),
                row.get("currency", ""),
                company_value(row),
                row.get("evidence_ref", ""),
                row.get("notes", ""),
            ]
        )

    rows.sort(key=lambda r: (priority.get(r[4], 99), r[3], r[2], r[0]))
    return header, rows


def build_purchase_type_rows(
    expenses: list[dict[str, str]],
    target_type: str,
) -> tuple[list[str], list[list[str]]]:
    header = [
        "entry_id",
        "date",
        "item",
        "initial_price_pkr",
        "current_amount_pkr",
        "status",
        "amount_status",
        "purchase_registration",
        "payment_status",
        "delivery_status",
        "procurement_stage",
        "company_name",
        "evidence_ref",
    ]

    rows: list[list[str]] = []
    for row in expenses:
        if classify_purchase_type(row) != target_type:
            continue

        amount = parse_decimal(row.get("amount", ""))
        status = (row.get("status") or "").strip().lower()
        amount_status = (row.get("amount_status") or "").strip().lower()

        initial_price = ""
        current_amount = ""
        if amount is not None and amount > 0 and amount_status == "confirmed":
            if status == "quote":
                initial_price = row.get("amount", "")
            else:
                current_amount = row.get("amount", "")

        rows.append(
            [
                row.get("entry_id", ""),
                row.get("date", ""),
                row.get("item", ""),
                initial_price,
                current_amount,
                row.get("status", ""),
                row.get("amount_status", ""),
                derive_purchase_registration(row),
                row.get("payment_status", ""),
                row.get("delivery_status", ""),
                row.get("procurement_stage", ""),
                company_value(row),
                row.get("evidence_ref", ""),
            ]
        )

    rows.sort(key=lambda r: (r[2], r[0]))
    return header, rows


def build_summary_sheet(workbook: Workbook, row_counts: dict[str, int], expenses: list[dict[str, str]]) -> None:
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "J40 Purchase Control"
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].fill = TITLE_FILL
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.merge_cells("A1:D1")

    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Workbook Path"
    summary["B4"] = str(OUTPUT_PATH)
    summary["A5"] = "Source of Truth"
    summary["B5"] = "CSV rows in data/manual/expenses.csv remain the editable source."
    for cell in ("A3", "A4", "A5"):
        summary[cell].font = Font(bold=True)

    section_row = 7
    summary[f"A{section_row}"] = "Sheet Coverage"
    summary[f"A{section_row}"].fill = SECTION_FILL
    summary[f"A{section_row}"].font = Font(bold=True)
    summary.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=4)

    summary.append(["Sheet", "Rows", "Purpose", "Status"])
    style_header_row(summary, section_row + 1)
    for sheet_name in ("Initial Price", "Purchase Registration", "Purchase of Goods", "Purchase of Services"):
        summary.append([sheet_name, row_counts[sheet_name], SHEET_NOTES[sheet_name], "Active"])

    next_row = summary.max_row + 2
    summary[f"A{next_row}"] = "Purchase Snapshot"
    summary[f"A{next_row}"].fill = SECTION_FILL
    summary[f"A{next_row}"].font = Font(bold=True)
    summary.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)

    summary.append(["Metric", "Amount / Count", "Scope", "Rule"])
    style_header_row(summary, next_row + 1)

    initial_price_total = Decimal("0")
    goods_registered_total = Decimal("0")
    services_registered_total = Decimal("0")
    registration_counts: Counter[str] = Counter()

    for row in expenses:
        amount = parse_decimal(row.get("amount", ""))
        amount_status = (row.get("amount_status") or "").strip().lower()
        status = (row.get("status") or "").strip().lower()
        purchase_type = classify_purchase_type(row)
        registration = derive_purchase_registration(row)
        registration_counts[registration] += 1

        if amount is not None and amount > 0 and amount_status == "confirmed" and status == "quote":
            initial_price_total += amount
        if amount is not None and amount_status == "confirmed" and status in {"paid", "received", "installed", "credited"}:
            if purchase_type == "goods":
                goods_registered_total += amount
            else:
                services_registered_total += amount

    summary_rows = [
        ["Captured initial price subtotal (quoted)", float(initial_price_total), "Initial Price", "Sum of positive confirmed quote rows"],
        ["Registered goods subtotal", float(goods_registered_total), "Purchase of Goods", "Sum of confirmed paid/received/installed/credited rows"],
        ["Registered services subtotal", float(services_registered_total), "Purchase of Services", "Sum of confirmed paid/received/installed/credited rows"],
        ["Rows needing registration confirmation", registration_counts["registered_needs_confirmation"], "Purchase Registration", "Rows marked registered but still missing proof alignment"],
    ]

    for metric, value, scope, rule in summary_rows:
        summary.append([metric, value, scope, rule])
        amount_cell = summary.cell(row=summary.max_row, column=2)
        if isinstance(amount_cell.value, (int, float)):
            amount_cell.number_format = '#,##0.00'

    summary.freeze_panes = "A8"
    autosize_sheet(summary)


def main() -> None:
    expenses_header, expenses_rows = read_csv(EXPENSES_PATH)
    expenses = [dict(zip(expenses_header, row)) for row in expenses_rows]

    initial_header, initial_rows = build_initial_price_rows(expenses)
    registration_header, registration_rows = build_purchase_registration_rows(expenses)
    goods_header, goods_rows = build_purchase_type_rows(expenses, "goods")
    services_header, services_rows = build_purchase_type_rows(expenses, "services")

    workbook = Workbook()
    row_counts = {
        "Initial Price": write_rows_sheet(workbook, "Initial Price", initial_header, initial_rows),
        "Purchase Registration": write_rows_sheet(workbook, "Purchase Registration", registration_header, registration_rows),
        "Purchase of Goods": write_rows_sheet(workbook, "Purchase of Goods", goods_header, goods_rows),
        "Purchase of Services": write_rows_sheet(workbook, "Purchase of Services", services_header, services_rows),
    }

    build_summary_sheet(workbook, row_counts, expenses)

    workbook._sheets = [workbook["Summary"]] + [sheet for sheet in workbook.worksheets if sheet.title != "Summary"]
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
