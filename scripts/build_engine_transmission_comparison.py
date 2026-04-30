from __future__ import annotations

import argparse
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parent.parent
DOCS_DIR = ROOT / "docs"
DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")
DEFAULT_REPORT_PATH = DOCS_DIR / "engine-transmission-cost-comparison.md"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

SOURCES = [
    "https://www.olx.com.pk/spare-parts_c82/q-engine-turbo?page=2",
    "https://www.olx.com.pk/garri-shahu_g5000685/q-engine",
    "https://www.olx.com.pk/nazir-garden-society_g5000674/engines_c709021/q-part-1",
    "https://www.olx.com.pk/punjab_g2003006/engines_c709021/q-1-",
    "https://www.olx.com.pk/islamabad_g4060615/spare-parts_c82/q-toyota-surf-hilux",
    "https://www.olx.com.pk/item/engine-3rz-atu-gear-iid-1111948382",
    "https://hdautomotive.com.au/product/brand-new-h55f-transmission-suitable-for-hzj70-hzj73-hzj75-series-landcruiser/",
    "https://landcruisercomponents.com/product/toyota-landcruiser-diesel-hzj75-1hz-h55-h55f-gearbox-transfer-recondition-outright-del-included/",
    "https://www.atfspeed.com/a340e-race-master-overhaul-kit.html",
    "https://bigcountrycustoms.com/product/r151f-kit-turbo/",
    "https://www.pakistanpoint.com/en/story/2165242/currency-rate-in-pakistan-dollar-euro-pound-riyal-rates-on-6-april-2026.html",
]


@dataclass(frozen=True)
class CostRange:
    min_pkr: int
    max_pkr: int

    @property
    def midpoint_pkr(self) -> int:
        return int((self.min_pkr + self.max_pkr) / 2)

    def add(self, other: "CostRange") -> "CostRange":
        return CostRange(self.min_pkr + other.min_pkr, self.max_pkr + other.max_pkr)

    def subtract(self, other: "CostRange") -> "CostRange":
        return CostRange(self.min_pkr - other.min_pkr, self.max_pkr - other.max_pkr)


@dataclass(frozen=True)
class TransmissionPath:
    name: str
    extra_cost: CostRange
    note: str
    confidence: str


@dataclass(frozen=True)
class EngineSwapOption:
    option: str
    engine_package: CostRange
    transmission_pairing: CostRange
    integration_and_fabrication: CostRange
    recommended_transmission: str
    confidence: str
    market_anchor: str

    def total_with_base(self, base: CostRange) -> CostRange:
        return base.add(self.engine_package).add(self.transmission_pairing).add(self.integration_and_fabrication)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").strip().split())


def parse_range_pkr(raw: object) -> CostRange:
    text = clean_text(raw)
    if not text:
        return CostRange(0, 0)
    if "-" in text:
        left, right = text.split("-", 1)
        left_digits = int("".join(ch for ch in left if ch.isdigit()) or 0)
        right_digits = int("".join(ch for ch in right if ch.isdigit()) or 0)
        return CostRange(left_digits, right_digits)
    digits = int("".join(ch for ch in text if ch.isdigit()) or 0)
    return CostRange(digits, digits)


def autosize_columns(ws: Worksheet, max_width: int = 70) -> None:
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        width = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=column).value
            if value is None:
                continue
            width = max(width, len(str(value)))
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def style_header_row(ws: Worksheet, row: int) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def format_pkr(value: int) -> str:
    return f"PKR {value:,}"


def format_range(value: CostRange) -> str:
    return f"{format_pkr(value.min_pkr)} to {format_pkr(value.max_pkr)}"


def get_baselines_from_parts_estimates(workbook_path: Path) -> tuple[CostRange, CostRange, CostRange]:
    workbook = load_workbook(workbook_path, data_only=True)
    if "Parts_Estimates" not in workbook.sheetnames:
        raise ValueError("`Parts_Estimates` tab not found in workbook.")

    ws = workbook["Parts_Estimates"]
    high_medium_rows: list[dict[str, object]] = []
    for row in range(2, ws.max_row + 1):
        category = clean_text(ws.cell(row=row, column=1).value)
        subcategory = clean_text(ws.cell(row=row, column=2).value)
        item = clean_text(ws.cell(row=row, column=3).value)
        priority = clean_text(ws.cell(row=row, column=8).value).lower()
        if not item or priority not in {"high", "medium"}:
            continue
        high_medium_rows.append(
            {
                "category": category,
                "subcategory": subcategory,
                "item": item,
                "range": parse_range_pkr(ws.cell(row=row, column=6).value),
            }
        )

    optional_keywords = ("old man emu suspension kit", "steering damper / stabilizer")
    baseline_with_eps_rows = [
        row for row in high_medium_rows if row["item"].lower() not in optional_keywords
    ]

    eps_core_subcategories = {"eps donor", "column hardware", "mounting", "seals / trim"}
    eps_core_rows = [
        row
        for row in baseline_with_eps_rows
        if row["category"] == "Steering / EPS" and row["subcategory"].lower() in eps_core_subcategories
    ]

    def sum_ranges(rows: list[dict[str, object]]) -> CostRange:
        min_total = sum(int(row["range"].min_pkr) for row in rows)
        max_total = sum(int(row["range"].max_pkr) for row in rows)
        return CostRange(min_total, max_total)

    baseline_with_eps = sum_ranges(baseline_with_eps_rows)
    eps_core = sum_ranges(eps_core_rows)
    baseline_manual_steering = baseline_with_eps.subtract(eps_core)
    return baseline_manual_steering, baseline_with_eps, eps_core


def build_transmission_paths() -> list[TransmissionPath]:
    return [
        TransmissionPath(
            name="Current gearbox service only (linkage, oil, minor shift hardware)",
            extra_cost=CostRange(30_000, 120_000),
            note="Low-cost first attempt to fix clunky shift feel.",
            confidence="Medium",
        ),
        TransmissionPath(
            name="Local rebuild of current gearbox + synchro/bearing work",
            extra_cost=CostRange(200_000, 500_000),
            note="Workshop overhaul path when service-level fixes are not enough.",
            confidence="Low-Medium",
        ),
        TransmissionPath(
            name="Used replacement gearbox from local market",
            extra_cost=CostRange(90_000, 250_000),
            note="Anchored by local OLX gearbox listings; add install and fluids.",
            confidence="Medium",
        ),
        TransmissionPath(
            name="Imported reconditioned H55F + transfer package",
            extra_cost=CostRange(1_100_000, 1_600_000),
            note="Uses US$3,278 package anchor plus freight/duty/install spread.",
            confidence="Low-Medium",
        ),
        TransmissionPath(
            name="Imported brand-new H55F",
            extra_cost=CostRange(1_200_000, 1_800_000),
            note="Uses AUD$4,970 anchor plus freight/duty/install spread.",
            confidence="Low-Medium",
        ),
    ]


def build_engine_swap_options() -> list[EngineSwapOption]:
    return [
        EngineSwapOption(
            option="Toyota 3B diesel + 4x4 gearbox package",
            engine_package=CostRange(350_000, 500_000),
            transmission_pairing=CostRange(0, 120_000),
            integration_and_fabrication=CostRange(180_000, 350_000),
            recommended_transmission="Use included 4x4 gearbox; service before install.",
            confidence="Medium",
            market_anchor="OLX listing around PKR 4.37 lac for 3B + 4x4 gearbox package.",
        ),
        EngineSwapOption(
            option="Toyota 1HZ 4.2 diesel + manual gearbox package",
            engine_package=CostRange(750_000, 1_050_000),
            transmission_pairing=CostRange(0, 150_000),
            integration_and_fabrication=CostRange(150_000, 350_000),
            recommended_transmission="Factory-style manual pairing; refresh clutch/hydraulics.",
            confidence="Medium",
            market_anchor="OLX listing around PKR 9 lac for 1HZ with manual gearbox.",
        ),
        EngineSwapOption(
            option="Toyota 1HD-FTE 4.2 turbo diesel",
            engine_package=CostRange(780_000, 1_200_000),
            transmission_pairing=CostRange(180_000, 450_000),
            integration_and_fabrication=CostRange(350_000, 800_000),
            recommended_transmission="Heavy-duty manual (H55/R151 family) with upgraded cooling and fuel support.",
            confidence="Low-Medium",
            market_anchor="OLX listing around PKR 8.6 lac for 1HD-FTE engine; integration is the expensive part.",
        ),
        EngineSwapOption(
            option="Toyota 1KZ-TE turbo diesel",
            engine_package=CostRange(350_000, 500_000),
            transmission_pairing=CostRange(120_000, 350_000),
            integration_and_fabrication=CostRange(250_000, 600_000),
            recommended_transmission="R-series manual or matched automatic donor setup.",
            confidence="Medium",
            market_anchor="OLX anchors around PKR 3.9-4.2 lac for 1KZ turbo engine.",
        ),
        EngineSwapOption(
            option="Toyota 1KD turbo intercooler",
            engine_package=CostRange(650_000, 900_000),
            transmission_pairing=CostRange(150_000, 400_000),
            integration_and_fabrication=CostRange(400_000, 900_000),
            recommended_transmission="Matched manual/auto donor transmission with ECU/immobilizer handling.",
            confidence="Low-Medium",
            market_anchor="OLX anchor around PKR 7.5 lac for 1KD turbo/intercooler setup.",
        ),
        EngineSwapOption(
            option="Toyota 3RZ-FE petrol",
            engine_package=CostRange(160_000, 350_000),
            transmission_pairing=CostRange(0, 220_000),
            integration_and_fabrication=CostRange(220_000, 500_000),
            recommended_transmission="Use complete engine+gear donor if possible to cut integration risk.",
            confidence="Medium",
            market_anchor="OLX anchors from PKR 1.6 to 3.5 lac (including some with auto gear).",
        ),
        EngineSwapOption(
            option="Toyota 1FZ-FE petrol (complete with auto gear)",
            engine_package=CostRange(480_000, 750_000),
            transmission_pairing=CostRange(0, 150_000),
            integration_and_fabrication=CostRange(260_000, 600_000),
            recommended_transmission="Use bundled auto/manual set and refresh mounts/cooling/exhaust.",
            confidence="Low-Medium",
            market_anchor="OLX anchor around PKR 6 lac for 1FZ-FE complete with auto gear.",
        ),
        EngineSwapOption(
            option="Toyota 1UZ-FE V8",
            engine_package=CostRange(150_000, 320_000),
            transmission_pairing=CostRange(200_000, 700_000),
            integration_and_fabrication=CostRange(450_000, 1_000_000),
            recommended_transmission="A340/A341 family with adapter package and full custom integration.",
            confidence="Low-Medium",
            market_anchor="OLX anchors around PKR 1.5 to 3.0 lac for 1UZ VVTi engines.",
        ),
    ]


def write_workbook_sheet(
    workbook_path: Path,
    baseline_manual: CostRange,
    baseline_eps: CostRange,
    eps_core: CostRange,
    transmission_paths: list[TransmissionPath],
    engine_options: list[EngineSwapOption],
) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = workbook_path.with_name(f"{workbook_path.stem}.engine_compare_backup_{timestamp}{workbook_path.suffix}")
    shutil.copy2(workbook_path, backup_path)

    workbook = load_workbook(workbook_path)
    sheet_name = "Engine_Transmission_Comparison"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)

    ws.append(["Engine / Transmission Cost Comparison"])
    ws.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    ws.append(["Workbook", str(workbook_path)])
    ws.append(["Baseline (manual steering, no OME, no steering damper)", format_range(baseline_manual)])
    ws.append(["EPS conversion delta", format_range(eps_core)])
    ws.append(["Baseline (with EPS, no OME, no steering damper)", format_range(baseline_eps)])
    ws.append(
        [
            "FX assumptions for imported gearbox references",
            "USD->PKR ~= 280, AUD->PKR ~= 194 (Pakistan open-market snapshot 6 Apr 2026)",
        ]
    )
    ws.append([])

    current_header_row = ws.max_row + 1
    ws.append(
        [
            "Scenario Group",
            "Option",
            "Base Build Cost",
            "Transmission Delta",
            "Total Build Cost",
            "Confidence",
            "Notes",
        ]
    )
    style_header_row(ws, current_header_row)

    for path in transmission_paths:
        total = baseline_eps.add(path.extra_cost)
        ws.append(
            [
                "Current engine retained",
                path.name,
                format_range(baseline_eps),
                format_range(path.extra_cost),
                format_range(total),
                path.confidence,
                path.note,
            ]
        )

    ws.append([])

    swap_header_row = ws.max_row + 1
    ws.append(
        [
            "Scenario Group",
            "Engine Option",
            "Engine Package",
            "Transmission Pairing",
            "Integration / Fabrication",
            "Total Build Cost (with EPS baseline)",
            "Recommended Transmission",
            "Confidence",
            "Market Anchor",
        ]
    )
    style_header_row(ws, swap_header_row)

    for option in engine_options:
        total = option.total_with_base(baseline_eps)
        ws.append(
            [
                "Engine replacement",
                option.option,
                format_range(option.engine_package),
                format_range(option.transmission_pairing),
                format_range(option.integration_and_fabrication),
                format_range(total),
                option.recommended_transmission,
                option.confidence,
                option.market_anchor,
            ]
        )

    ws.append([])
    source_header_row = ws.max_row + 1
    ws.append(["Sources used"])
    ws.cell(row=source_header_row, column=1).font = HEADER_FONT
    ws.cell(row=source_header_row, column=1).fill = HEADER_FILL
    for source in SOURCES:
        ws.append([source])

    autosize_columns(ws)
    workbook.save(workbook_path)
    return backup_path


def build_markdown_report(
    report_path: Path,
    workbook_path: Path,
    baseline_manual: CostRange,
    baseline_eps: CostRange,
    eps_core: CostRange,
    transmission_paths: list[TransmissionPath],
    engine_options: list[EngineSwapOption],
) -> None:
    lines: list[str] = []
    lines.append("# Engine + Transmission Cost Comparison")
    lines.append("")
    lines.append(f"- Generated: `{datetime.now().isoformat(timespec='seconds')}`")
    lines.append(f"- Workbook: `{workbook_path}`")
    lines.append(f"- Baseline (manual steering, no OME/damper): `{format_range(baseline_manual)}`")
    lines.append(f"- EPS conversion delta: `{format_range(eps_core)}`")
    lines.append(f"- Baseline with EPS (no OME/damper): `{format_range(baseline_eps)}`")
    lines.append("- FX assumptions for imported gearbox references: `USD->PKR ~= 280`, `AUD->PKR ~= 194`.")
    lines.append("")
    lines.append("## Current Engine Retained (with EPS baseline)")
    lines.append("")
    lines.append("| Option | Transmission Delta | Total Build Cost | Confidence |")
    lines.append("|---|---:|---:|---|")
    for path in transmission_paths:
        total = baseline_eps.add(path.extra_cost)
        lines.append(
            f"| {path.name} | {format_range(path.extra_cost)} | {format_range(total)} | {path.confidence} |"
        )
    lines.append("")
    lines.append("## Engine Replacement Options (with EPS baseline)")
    lines.append("")
    lines.append("| Option | Engine Package | Transmission Pairing | Integration / Fabrication | Total Build Cost | Confidence |")
    lines.append("|---|---:|---:|---:|---:|---|")
    for option in engine_options:
        total = option.total_with_base(baseline_eps)
        lines.append(
            f"| {option.option} | {format_range(option.engine_package)} | {format_range(option.transmission_pairing)} | "
            f"{format_range(option.integration_and_fabrication)} | {format_range(total)} | {option.confidence} |"
        )
    lines.append("")
    lines.append("## Notes")
    lines.append("")
    lines.append("- These are planning ranges, not quotations.")
    lines.append("- Engine package prices are listing anchors from OLX and can move quickly by condition and location.")
    lines.append("- Integration/fabrication is the biggest uncertainty driver (mounts, wiring/ECU, exhaust, cooling, driveshaft work).")
    lines.append("- Most accurate next step: collect 3 written quotes per shortlisted engine path and replace range assumptions with quote values.")
    lines.append("")
    lines.append("## Source URLs")
    lines.append("")
    for source in SOURCES:
        lines.append(f"- {source}")
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build engine/transmission scenario comparison and write into workbook.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH, help="Workbook path to update in-place.")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH, help="Markdown report output path.")
    args = parser.parse_args()

    baseline_manual, baseline_eps, eps_core = get_baselines_from_parts_estimates(args.workbook)
    transmission_paths = build_transmission_paths()
    engine_options = build_engine_swap_options()

    backup_path = write_workbook_sheet(
        workbook_path=args.workbook,
        baseline_manual=baseline_manual,
        baseline_eps=baseline_eps,
        eps_core=eps_core,
        transmission_paths=transmission_paths,
        engine_options=engine_options,
    )
    build_markdown_report(
        report_path=args.report,
        workbook_path=args.workbook,
        baseline_manual=baseline_manual,
        baseline_eps=baseline_eps,
        eps_core=eps_core,
        transmission_paths=transmission_paths,
        engine_options=engine_options,
    )

    print(f"Updated workbook: {args.workbook}")
    print(f"Backup saved: {backup_path}")
    print(f"Report written: {args.report}")
    print(f"Baseline manual steering: {format_range(baseline_manual)}")
    print(f"Baseline with EPS: {format_range(baseline_eps)}")


if __name__ == "__main__":
    main()
