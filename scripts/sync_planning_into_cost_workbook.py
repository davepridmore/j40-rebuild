from __future__ import annotations

import argparse
import csv
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parent.parent
MANUAL_DIR = ROOT / "data" / "manual"

DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


SHEET_INPUTS: tuple[tuple[str, Path], ...] = (
    ("Plan_Summary", MANUAL_DIR / "procurement_decision_matrix_pass2.csv"),
    ("Procurement_Pass2", MANUAL_DIR / "procurement_decision_matrix_pass2.csv"),
    ("Procurement_Baskets", MANUAL_DIR / "procurement_local_baskets_pass2.csv"),
    ("Reassembly_Packages", MANUAL_DIR / "reassembly_work_packages.csv"),
    ("Reassembly_Dependency", MANUAL_DIR / "reassembly_dependency_edges.csv"),
    ("Component_Disposition", MANUAL_DIR / "component_disposition_plan.csv"),
    ("Procurement_Base", MANUAL_DIR / "procurement_decision_matrix.csv"),
    ("Cost_Reconciliation", MANUAL_DIR / "j40_costs_expenses_reconciliation.csv"),
    ("Workbook_Tidy_Backlog", MANUAL_DIR / "j40_workbook_tidy_backlog.csv"),
)


def load_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        return fieldnames, list(reader)


def autosize_columns(ws: Worksheet, max_width: int = 65) -> None:
    for col_idx in range(1, ws.max_column + 1):
        column = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[column].width = min(max(max_len + 2, 10), max_width)


def write_table_sheet(workbook, sheet_name: str, csv_path: Path) -> tuple[int, int]:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(title=sheet_name)

    fieldnames, rows = load_csv_rows(csv_path)
    if not fieldnames:
        ws["A1"] = "No data"
        return 0, 0

    for col_idx, header in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(fieldnames, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{max(len(rows) + 1, 1)}"
    autosize_columns(ws)
    return len(rows), len(fieldnames)


def summarize_pass2(pass2_rows: list[dict[str, str]]) -> dict[str, object]:
    decision_counts: dict[str, int] = {}
    timing_counts: dict[str, int] = {}
    immediate_rows: list[dict[str, str]] = []

    for row in pass2_rows:
        decision = row.get("pass2_decision", "")
        timing = row.get("timing_window", "")
        decision_counts[decision] = decision_counts.get(decision, 0) + 1
        timing_counts[timing] = timing_counts.get(timing, 0) + 1
        if timing in {"tub_off_immediate", "in_flight_now"}:
            immediate_rows.append(row)

    return {
        "decision_counts": decision_counts,
        "timing_counts": timing_counts,
        "immediate_rows": immediate_rows,
    }


def write_plan_summary_sheet(workbook, summary_csv_path: Path) -> None:
    if "Plan_Summary" in workbook.sheetnames:
        del workbook["Plan_Summary"]
    ws = workbook.create_sheet(title="Plan_Summary")

    _, rows = load_csv_rows(summary_csv_path)
    summary = summarize_pass2(rows)

    ws["A1"] = "Codex Reconciliation Summary"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Source Workbook"
    ws["B4"] = str(DEFAULT_WORKBOOK_PATH)
    ws["A5"] = "Rows Evaluated"
    ws["B5"] = len(rows)

    ws["A7"] = "Pass2 Decision Counts"
    ws["A7"].font = Font(bold=True)
    row = 8
    for key, value in sorted(summary["decision_counts"].items()):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws["D7"] = "Timing Window Counts"
    ws["D7"].font = Font(bold=True)
    row = 8
    for key, value in sorted(summary["timing_counts"].items()):
        ws.cell(row=row, column=4, value=key)
        ws.cell(row=row, column=5, value=value)
        row += 1

    ws["A20"] = "Immediate Actions (Tub-Off / In-Flight)"
    ws["A20"].font = Font(bold=True)
    ws["A21"] = "entry_id"
    ws["B21"] = "item"
    ws["C21"] = "pass2_decision"
    ws["D21"] = "timing_window"
    for col in ("A21", "B21", "C21", "D21"):
        ws[col].fill = HEADER_FILL
        ws[col].font = HEADER_FONT

    row = 22
    for item in summary["immediate_rows"]:
        ws.cell(row=row, column=1, value=item.get("entry_id", ""))
        ws.cell(row=row, column=2, value=item.get("item", ""))
        ws.cell(row=row, column=3, value=item.get("pass2_decision", ""))
        ws.cell(row=row, column=4, value=item.get("timing_window", ""))
        row += 1

    ws.freeze_panes = "A22"
    autosize_columns(ws, max_width=80)


def backup_workbook(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = path.with_name(f"{path.stem}.backup_{stamp}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync Codex planning outputs into J40_Costs workbook.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to workbook to update in-place.",
    )
    parser.add_argument(
        "--skip-backup",
        action="store_true",
        help="Skip backup creation before update.",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    missing_inputs = [str(path) for _, path in SHEET_INPUTS if not path.exists()]
    if missing_inputs:
        raise FileNotFoundError(f"Missing input CSV files: {', '.join(missing_inputs)}")

    backup_path = None
    if not args.skip_backup:
        backup_path = backup_workbook(workbook_path)

    workbook = load_workbook(workbook_path)

    write_plan_summary_sheet(workbook, MANUAL_DIR / "procurement_decision_matrix_pass2.csv")

    sheet_stats: list[tuple[str, int, int]] = []
    for sheet_name, csv_path in SHEET_INPUTS:
        if sheet_name == "Plan_Summary":
            continue
        rows, columns = write_table_sheet(workbook, sheet_name, csv_path)
        sheet_stats.append((sheet_name, rows, columns))

    workbook.save(workbook_path)

    print(f"Updated workbook: {workbook_path}")
    if backup_path:
        print(f"Backup created: {backup_path}")
    print("Managed sheets updated:")
    for name, rows, columns in sheet_stats:
        print(f"- {name}: {rows} rows, {columns} columns")


if __name__ == "__main__":
    main()
