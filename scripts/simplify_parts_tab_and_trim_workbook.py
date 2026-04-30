from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")

ACTIVE_DECISIONS = {
    "buy_minimum_qty_now",
    "measure_then_buy_with_trial_fit",
    "grade_spec_then_buy",
    "prepare_before_trial_fit",
    "measure_then_order_local",
    "lock_vendor_then_buy_for_reassembly",
    "bundle_local_toyota_buy_after_inspection",
    "inspect_then_local_decide",
    "track_in_flight_order",
}

LEGACY_PLANNED_ITEMS = [
    ("Carpets/Mats", "post_baseline_only", "defer_until_scope_review", "Legacy planned interior item retained for visibility."),
    ("shackle reversal", "post_baseline_only", "defer_until_scope_review", "Legacy planned suspension option retained for visibility."),
    ("Electrical power steering", "pre_tub_refit", "plan_and_buy_when_ready", "Legacy planned item retained; EPS path remains in scope."),
    ("Steering bushings", "post_tub_off_inspection", "inspect_then_local_decide", "Legacy planned steering service item retained for visibility."),
    ("Hidden evaporator/heater/defrost unit, compact Mini-size", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Slim under-dash louver outlet panel", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Separate 3-knob control panel", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("2.5-inch duct hose and defrost hose kit", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Drain hose and mounting kit", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Parallel-flow condenser sized to your core support", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Receiver-drier", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Trinary switch", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Barrier hose and fittings", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Firewall bulkhead fittings", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("Relay/fuse/wiring for blower, clutch, and fan", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC stack item retained for visibility."),
    ("New O-rings, refrigerant oil, and full R134a charge setup", "post_baseline_only", "defer_until_scope_review", "Legacy planned HVAC consumables retained for visibility."),
]


def normalize(text: object) -> str:
    if text is None:
        return ""
    return " ".join(str(text).replace("\n", " ").strip().lower().split())


def collect_active_procurement(workbook) -> list[dict[str, str]]:
    ws = workbook["Procurement_Pass2"]
    items: list[dict[str, str]] = []
    for row in range(2, ws.max_row + 1):
        entry_id = ws.cell(row=row, column=1).value
        item = ws.cell(row=row, column=3).value
        timing = ws.cell(row=row, column=6).value
        decision = ws.cell(row=row, column=7).value
        supplier_hint = ws.cell(row=row, column=10).value
        rationale = ws.cell(row=row, column=11).value

        if not item:
            continue
        decision_token = normalize(decision)
        if decision_token not in ACTIVE_DECISIONS:
            continue

        items.append(
            {
                "entry_id": str(entry_id or ""),
                "item": str(item).strip(),
                "timing": str(timing or "").strip(),
                "decision": str(decision or "").strip(),
                "supplier_hint": str(supplier_hint or "").strip(),
                "rationale": str(rationale or "").strip(),
            }
        )
    return items


def simplify_parts_tab(workbook, active_items: list[dict[str, str]]) -> tuple[int, int]:
    ws = workbook["Parts"]

    # Make planning columns explicit so the main tab stays readable.
    ws.cell(row=1, column=6).value = "Plan_Timing"
    ws.cell(row=1, column=7).value = "Plan_Decision"

    # Keep only meaningful existing rows: drop old section headers and previously generated active-plan rows.
    kept_rows: list[list[object]] = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, 14)]
        if not any(v not in (None, "") for v in values):
            continue
        status = normalize(values[10] if len(values) > 10 else "")
        if status == "section_header":
            continue
        note = normalize(values[11] if len(values) > 11 else "")
        generated_note_prefixes = (
            "active required purchase item from procurement_pass2.",
            "condition-based item: inspect first, then buy only if required.",
            "already ordered/in-flight in procurement and mirrored by wiring grommet entry; avoid duplicate buy.",
        )
        if any(note.startswith(prefix) for prefix in generated_note_prefixes):
            continue
        kept_rows.append(values)

    existing_names = {normalize(row[0]) for row in kept_rows if row[0]}

    # Add one compact separator for the active plan.
    kept_rows.append(
        [
            "Required Purchases (Active Plan)",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "section_header",
            "Planning section for active required buys only.",
            None,
        ]
    )

    added = 0
    for item in active_items:
        key = normalize(item["item"])
        if key in existing_names:
            continue

        if item["entry_id"] == "part_rubber_grommet_set":
            reconciled_status = "needs_confirmation"
            reconciled_note = (
                "Already ordered/in-flight in procurement and mirrored by Wiring grommet entry; avoid duplicate buy."
            )
            procured_mark = "ordered/in-flight"
            received = "Needs check"
            paid = "9200 COD (ref Wiring)"
        else:
            decision_key = normalize(item["decision"])
            if decision_key == "inspect_then_local_decide":
                reconciled_status = "planned_or_open"
                reconciled_note = "Condition-based item: inspect first, then buy only if required."
            else:
                reconciled_status = "planned_or_open"
                reconciled_note = "Active required purchase item from Procurement_Pass2."
            procured_mark = None
            received = None
            paid = None

        kept_rows.append(
            [
                item["item"],
                None,
                item["supplier_hint"] or None,
                received,
                paid,
                item["timing"] or None,
                item["decision"] or None,
                None,
                None,
                None,
                reconciled_status,
                reconciled_note,
                procured_mark,
            ]
        )
        existing_names.add(key)
        added += 1

    # Retain previously discussed planned items as explicit rows (not empty headers).
    kept_rows.append(
        [
            "Legacy Planned Items (Retained)",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "section_header",
            "Previously discussed items retained for visibility even when deferred.",
            None,
        ]
    )
    for item_name, timing, decision, note in LEGACY_PLANNED_ITEMS:
        key = normalize(item_name)
        if key in existing_names:
            continue
        kept_rows.append(
            [
                item_name,
                None,
                None,
                None,
                None,
                timing,
                decision,
                None,
                None,
                None,
                "planned_or_open",
                note,
                None,
            ]
        )
        existing_names.add(key)

    # Rewrite rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, 14):
            ws.cell(row=row, column=col).value = None

    for idx, row_values in enumerate(kept_rows, start=2):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=idx, column=col).value = value

    return len(kept_rows), added


def main() -> None:
    parser = argparse.ArgumentParser(description="Simplify workbook tabs and make Parts tab the active required-purchase list.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH, help="Workbook path to update in-place.")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = args.workbook.with_name(f"{args.workbook.stem}.simplify_parts_backup_{timestamp}{args.workbook.suffix}")
    shutil.copy2(args.workbook, backup_path)

    workbook = load_workbook(args.workbook)
    active_items = collect_active_procurement(workbook)
    rows_after, added = simplify_parts_tab(workbook, active_items)

    # Remove non-essential planning tab requested by user.
    if "Supplier_Starting_Points" in workbook.sheetnames:
        del workbook["Supplier_Starting_Points"]

    workbook.save(args.workbook)

    print(f"Updated workbook: {args.workbook}")
    print(f"Backup saved: {backup_path}")
    print(f"Active procurement items considered: {len(active_items)}")
    print(f"Rows in Parts after simplification: {rows_after}")
    print(f"New required-purchase rows added to Parts: {added}")
    print("Removed tab: Supplier_Starting_Points")


if __name__ == "__main__":
    main()
