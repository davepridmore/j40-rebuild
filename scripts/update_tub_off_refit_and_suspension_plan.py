from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_WORKBOOK_PATH = Path("/Users/davidpridmore/Documents/J40_Costs.xlsx")
DEFAULT_REPORT_PATH = Path("docs/tub-off-refit-execution-plan.md")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def autosize_columns(ws: Worksheet, max_width: int = 70) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            width = max(width, len(str(value)))
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def style_header_row(ws: Worksheet, row_idx: int) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def update_build_plan(workbook) -> None:
    ws = workbook["Build_Plan"]
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if any(v not in (None, "") for v in vals):
            rows.append(vals)

    by_id = {str(r[1]): r for r in rows if r[1]}

    def upsert(record: list[object]) -> None:
        by_id[str(record[1])] = record

    upsert(
        [
            "work_package",
            "WP01A",
            "Tub Lift + Mount-Point Mapping",
            "body_structure",
            "queued",
            "stripdown_cataloguing_complete",
            "Tub lifted and every body/chassis mount point mapped before welding closure.",
            "Tag all mount points, captive nuts, shim positions, hole condition, and required repair action before refit planning.",
        ]
    )
    upsert(
        [
            "work_package",
            "WP01B",
            "Chassis/Tub Interface Weld Closure",
            "body_structure",
            "queued",
            "WP01A",
            "All mount interfaces weld-repaired, protected, and dimension-checked for refit.",
            "No final tub refit until all mount points pass straightness/fit and anti-corrosion prep checks.",
        ]
    )
    upsert(
        [
            "work_package",
            "WP02A",
            "Tub Refit Interface + Rubber Kit Control",
            "body_weather_seal",
            "queued",
            "WP01B",
            "Body mount rubbers, hardware, and shims selected and trial-fitted before final torque.",
            "Capture exact attachment points and shim stack plan before permanent tub-to-chassis fastening.",
        ]
    )
    upsert(
        [
            "work_package",
            "WP04B",
            "Tub-Off Engine Access Service + Inspection",
            "mechanical",
            "queued",
            "WP01A",
            "Engine-bay access inspection closed with condition-based replacement list.",
            "While tub is off: inspect mounts, hoses, lines, clamps, seals, linkage, and gearbox top-side service items.",
        ]
    )
    upsert(
        [
            "work_package",
            "WP04C",
            "Suspension Setup: OME Dampers + Local Springs",
            "steering_brakes_suspension",
            "queued",
            "WP01B",
            "Local front/rear leaf packs and OME dampers installed and aligned.",
            "Use local leaf springs for cost and serviceability; retain OME damping and bushing/hardware refresh.",
        ]
    )

    # Update WP06 dependency gate if present.
    if "WP06" in by_id:
        wp06 = by_id["WP06"]
        wp06[5] = "WP02|WP02A|WP03|WP04|WP04A|WP04B|WP04C|WP05"
        wp06[7] = (
            "Only release optional upgrades after baseline validation sign-off. "
            "Tub refit control and suspension setup must close first."
        )

    updated_rows = list(by_id.values())
    updated_rows.sort(key=lambda r: (0 if str(r[0]) == "work_package" else 1, str(r[1])))

    # Clear and rewrite.
    for r in range(2, ws.max_row + 1):
        for c in range(1, 9):
            ws.cell(r, c).value = None
    for idx, row in enumerate(updated_rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(idx, c).value = value

    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h


def update_procurement_pass2(workbook) -> None:
    ws = workbook["Procurement_Pass2"]
    headers = [ws.cell(1, c).value for c in range(1, 12)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 12)]
        if any(v not in (None, "") for v in vals):
            rows.append(vals)

    by_id = {str(r[0]): r for r in rows if r[0]}

    def upsert(record: list[object]) -> None:
        by_id[str(record[0])] = record

    # Activate OME shock path.
    upsert(
        [
            "part_old_man_emu_shocks",
            "steering_brakes_suspension",
            "Old Man Emu Nitrocharger shocks (front + rear)",
            "research_compare_then_select",
            "local_4x4_or_arb",
            "pre_tub_refit",
            "lock_vendor_then_buy_for_reassembly",
            "planned_upgrade_now",
            "basket_suspension_refit_bundle",
            "Use ARB/OME dealer or trusted local 4x4 suspension source.",
            "User decision: include OME dampers in current build phase before tub refit.",
        ]
    )

    # New local spring + tub mount interface rows.
    upsert(
        [
            "part_local_leaf_springs_front",
            "steering_brakes_suspension",
            "Local front leaf spring pack",
            "new_item",
            "local_suspension_workshop",
            "pre_tub_refit",
            "measure_then_order_local",
            "local_bundle_buy",
            "basket_suspension_refit_bundle",
            "Use local spring maker with load/ride-height brief and bushing compatibility check.",
            "Chosen for cost control and local serviceability while keeping OME dampers.",
        ]
    )
    upsert(
        [
            "part_local_leaf_springs_rear",
            "steering_brakes_suspension",
            "Local rear leaf spring pack",
            "new_item",
            "local_suspension_workshop",
            "pre_tub_refit",
            "measure_then_order_local",
            "local_bundle_buy",
            "basket_suspension_refit_bundle",
            "Use local spring maker with load/ride-height brief and anti-inversion compatibility check.",
            "Chosen for cost control and local serviceability while keeping OME dampers.",
        ]
    )
    upsert(
        [
            "part_body_mount_rubber_kit",
            "body_chassis",
            "Body-to-chassis mount rubber kit",
            "new_item",
            "local_rubber_or_landcruiser_supplier",
            "pre_tub_refit",
            "measure_then_buy_with_trial_fit",
            "critical_refit_buy",
            "basket_tub_refit_interface",
            "Local rubber/4x4 supplier or custom rubber fabricator if exact kit unavailable.",
            "Tub must not be bolted down until rubbers are test-fitted against mapped mount points.",
        ]
    )
    upsert(
        [
            "part_body_mount_hardware_kit",
            "body_chassis",
            "Body mount bolts, washers, sleeves, captive-nut repairs",
            "new_item",
            "local_fastener_and_fabrication",
            "pre_tub_refit",
            "grade_spec_then_buy",
            "critical_refit_buy",
            "basket_tub_refit_interface",
            "Use graded fasteners and include sleeve/spacer and captive-nut repair provision.",
            "Refit reliability depends on correct hardware grade and mount stack geometry.",
        ]
    )
    upsert(
        [
            "part_body_mount_shim_pack",
            "body_chassis",
            "Body mount shims/spacers alignment pack",
            "new_item",
            "local_fabrication",
            "pre_tub_refit",
            "prepare_before_trial_fit",
            "critical_refit_buy",
            "basket_tub_refit_interface",
            "Fabricate shim pack after mount-point map and weld closure dimensions are confirmed.",
            "Required to align tub seating and panel gaps before final torque.",
        ]
    )
    upsert(
        [
            "part_gearbox_top_service_items",
            "mechanical_baseline",
            "Gearbox top-cover service items (detents, bushes, shift-seat kit)",
            "new_item",
            "local_transmission_supplier",
            "post_tub_off_inspection",
            "inspect_then_local_decide",
            "inspect_first",
            "basket_condition_based_after_inspection",
            "Local transmission parts source; buy only against diagnosed wear.",
            "Targets disengagement/long-throw issue without committing to full rebuild.",
        ]
    )

    updated_rows = list(by_id.values())
    updated_rows.sort(key=lambda r: str(r[0]))

    for r in range(2, ws.max_row + 1):
        for c in range(1, 12):
            ws.cell(r, c).value = None
    for idx, row in enumerate(updated_rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(idx, c).value = value

    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h


def update_parts_estimates(workbook) -> None:
    ws = workbook["Parts_Estimates"]
    header = [ws.cell(1, c).value for c in range(1, 9)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if any(v not in (None, "") for v in vals):
            rows.append(vals)

    # Remove old all-in OME kit line to avoid double counting.
    filtered = []
    for row in rows:
        item = str(row[2] or "").strip().lower()
        if item == "old man emu suspension kit":
            continue
        filtered.append(row)
    rows = filtered

    def upsert_by_item(new_row: list[object]) -> None:
        target = str(new_row[2]).strip().lower()
        for i, row in enumerate(rows):
            if str(row[2] or "").strip().lower() == target:
                rows[i] = new_row
                return
        # Insert near suspension section.
        insert_at = None
        for i, row in enumerate(rows):
            if str(row[0] or "").strip() == "Suspension":
                insert_at = i + 1
        if insert_at is None:
            rows.append(new_row)
        else:
            rows.insert(insert_at, new_row)

    upsert_by_item(
        [
            "Suspension",
            "Dampers",
            "Old Man Emu Nitrocharger shocks (front + rear set)",
            "1 set",
            "ARB / OME supplier",
            "110000-220000",
            "170000",
            "High",
        ]
    )
    upsert_by_item(
        [
            "Suspension",
            "Leaf Springs",
            "Local fabricated leaf springs (front + rear)",
            "1 set",
            "Local spring maker / suspension workshop",
            "70000-180000",
            "120000",
            "High",
        ]
    )
    upsert_by_item(
        [
            "Suspension",
            "Setup",
            "Spring setup, shims, and alignment labor",
            "1 lot",
            "Local suspension workshop",
            "30000-90000",
            "60000",
            "Medium",
        ]
    )
    upsert_by_item(
        [
            "Body / Chassis",
            "Mounts",
            "Body mount shim/spacer set",
            "1 set",
            "Local fabrication / fastener supplier",
            "5000-20000",
            "15000",
            "High",
        ]
    )

    # Rewrite sheet.
    for r in range(2, ws.max_row + 1):
        for c in range(1, 9):
            ws.cell(r, c).value = None
    for idx, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(idx, c).value = value
    for c, h in enumerate(header, start=1):
        ws.cell(1, c).value = h


def update_suspension_sheet(workbook) -> None:
    ws = workbook["Suspension"]
    ws.cell(2, 1).value = "Front Leaf Springs"
    ws.cell(2, 2).value = "LOCAL-FJ40-FRONT-LEAF"
    ws.cell(2, 3).value = 2
    ws.cell(2, 4).value = "Local fabricated pack; tune ride height after tub refit trial fit."

    ws.cell(3, 1).value = "Rear Leaf Springs"
    ws.cell(3, 2).value = "LOCAL-FJ40-REAR-LEAF"
    ws.cell(3, 3).value = 2
    ws.cell(3, 4).value = "Local fabricated pack; tuned for load and anti-wrap behavior."

    ws.cell(4, 1).value = "Front Shocks"
    ws.cell(4, 2).value = "OME-60097"
    ws.cell(4, 3).value = 2
    ws.cell(4, 4).value = "Old Man Emu Nitrocharger Sport front shocks."

    ws.cell(5, 1).value = "Rear Shocks"
    ws.cell(5, 2).value = "OME-63064"
    ws.cell(5, 3).value = 2
    ws.cell(5, 4).value = "Old Man Emu Nitrocharger Sport rear shocks."


def write_tub_off_refit_plan_sheet(workbook) -> None:
    sheet_name = "Tub_Off_Refit_Plan"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)

    ws.append(["Tub-Off To Refit Control Plan"])
    ws.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    ws.append(["Intent", "Control welding + engine-access service + correct tub reattachment + OME dampers with local spring packs."])
    ws.append([])
    ws.append(
        [
            "phase_id",
            "lane",
            "task",
            "depends_on",
            "inspection_or_measurement",
            "parts_or_materials",
            "decision_gate",
            "completion_signal",
        ]
    )
    style_header_row(ws, 5)

    rows = [
        [
            "TO1",
            "body_structure",
            "Pre-lift baseline capture and tagging",
            "stripdown_cataloguing_complete",
            "Photo and tag every mount point and spacer before separation.",
            "Tagging consumables, marker labels, template sheet",
            "No tub lift until tag-map is complete.",
            "Mount map and reference photos signed off.",
        ],
        [
            "TO2",
            "body_structure",
            "Tub lift and mount-point condition survey",
            "TO1",
            "Record hole ovality, captive-nut condition, and corrosion depth at each point.",
            "Inspection tools, rust-penetrant, borescope as needed",
            "No welding closure without per-point repair decision.",
            "Mount condition matrix completed.",
        ],
        [
            "TO3",
            "body_structure",
            "Welding and closure of mount interfaces",
            "TO2",
            "Confirm mount faces, hole centerlines, and plate thickness restoration.",
            "Repair plates, primer, metal protection",
            "No refit planning until weld closure and anti-corrosion prep are complete.",
            "All mount points dimension-checked and protected.",
        ],
        [
            "TO4",
            "mechanical",
            "Tub-off engine-bay access inspection and service list lock",
            "TO2",
            "Inspect mounts, lines, hoses, clamps, shift linkage/top service items.",
            "Service parts bundle + condition-based items",
            "Buy only confirmed-failed items for condition-based components.",
            "Inspection report with buy/no-buy per line item.",
        ],
        [
            "TO5",
            "steering_brakes_suspension",
            "Suspension install path: OME dampers + local leaf packs",
            "TO3",
            "Validate spring arch, shackle angle, pinion/caster behavior after trial load.",
            "OME shocks, local front/rear springs, bushings, U-bolts, shims",
            "Do not final-torque suspension until ride-height and alignment checks pass.",
            "Suspension geometry check signed off.",
        ],
        [
            "TO6",
            "body_weather_seal",
            "Tub reattachment interface prep and trial fit",
            "TO3",
            "Dry trial fit with shim pack and rubber mounts before final bolt-up.",
            "Body mount rubber kit, hardware kit, shim/spacer pack",
            "No permanent tub fastening before successful trial seating.",
            "All attachment points align without forced fit.",
        ],
        [
            "TO7",
            "integration",
            "Final tub reattachment and torque map execution",
            "TO6|TO5|TO4",
            "Follow cross-pattern torque sequence and recheck after settling.",
            "Final hardware and thread treatment",
            "Gate to integration close only after torque recheck and gap alignment.",
            "Tub fixed at all correct points with verified torque log.",
        ],
        [
            "TO8",
            "integration",
            "Post-refit validation",
            "TO7",
            "Road/yard check for shift engagement, vibration, mount settling, and steering feel.",
            "None beyond completed assemblies",
            "Escalate only if disengagement or mount-settle issues remain.",
            "Vehicle moves to reassembly validation lane.",
        ],
    ]
    for row in rows:
        ws.append(row)

    autosize_columns(ws)


def write_report(path: Path) -> None:
    content = """# Tub-Off To Refit Execution Plan

- Keep welding scope separate, but lock interface control now.
- During tub-off, run a focused engine-access inspection and buy condition-based parts only.
- Before tub goes back, complete mount-point mapping, mount repairs, and trial fit using correct rubbers/hardware/shims.
- Suspension path is now fixed to: OME dampers plus local front/rear leaf spring packs.

## Key Procurement Adds

- Body-to-chassis mount rubber kit
- Body mount hardware and captive-nut repair provision
- Body mount shim/spacer alignment pack
- Local front and rear leaf spring packs
- OME Nitrocharger shock set (activated from deferred state)
- Gearbox top-cover service items (condition-based)

## Gates

1. No tub lift without mount-point tag map.
2. No refit plan without weld/dimension closure.
3. No final tub fastening without dry trial fit using rubbers and shim plan.
4. No suspension final torque until geometry/ride-height checks pass.
"""
    path.write_text(content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Update workbook with tub-off/refit control plan and suspension decision.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH, help="Workbook path to update in-place.")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH, help="Markdown report output path.")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = args.workbook.with_name(f"{args.workbook.stem}.tub_refit_plan_backup_{timestamp}{args.workbook.suffix}")
    shutil.copy2(args.workbook, backup)

    wb = load_workbook(args.workbook)
    update_build_plan(wb)
    update_procurement_pass2(wb)
    update_parts_estimates(wb)
    update_suspension_sheet(wb)
    write_tub_off_refit_plan_sheet(wb)
    wb.save(args.workbook)

    args.report.parent.mkdir(parents=True, exist_ok=True)
    write_report(args.report)

    print(f"Updated workbook: {args.workbook}")
    print(f"Backup saved: {backup}")
    print(f"Report written: {args.report}")


if __name__ == "__main__":
    main()
